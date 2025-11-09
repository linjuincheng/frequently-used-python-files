import os
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

# 設定資料夾路徑
folder_path = r'C:\_python 2024\2024final'

# 定義字體、對齊和邊框樣式
font = Font(name='新細明體', size=12)
alignment = Alignment(horizontal='center', vertical='center')
border = Border(left=Side(border_style='thin'), 
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin'))

def utc_to_taiwan(utc_str):
    """將 UTC 時間轉換為台灣時間"""
    try:
        utc_time = datetime.strptime(utc_str, '%Y-%m-%d %H:%M:%S.%f')
    except ValueError:
        try:
            utc_time = datetime.strptime(utc_str, '%Y-%m-%d %H:%M:%S')
        except ValueError:
            raise ValueError(f"Unsupported time format: {utc_str}")
    taiwan_time = utc_time + timedelta(hours=8)
    return taiwan_time.strftime('%Y-%m-%d %H:%M:%S')

def auto_adjust(ws):
    """自動調整欄寬與列高"""
    # 調整欄寬
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                try:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
                except:
                    pass
        ws.column_dimensions[col_letter].width = max_length + 2
    # 調整列高
    for row in ws.iter_rows():
        max_height = 0
        for cell in row:
            if cell.value:
                height = len(str(cell.value).split('\n'))
                if height > max_height:
                    max_height = height
        ws.row_dimensions[row[0].row].height = (max_height + 2) * 15

def apply_formatting(ws):
    """設定儲存格字型、對齊及邊框，並自動調整欄寬與列高"""
    for row in ws.iter_rows():
        for cell in row:
            cell.font = font
            cell.alignment = alignment
            cell.border = border
    auto_adjust(ws)

def process_excel(file_path):
    wb = load_workbook(file_path)

    # ---------------------------
    # Step 0: 讀取原始資料表 (Sheet1 或 sheet1)
    if 'Sheet1' in wb.sheetnames:
        sheet1 = wb['Sheet1']
    elif 'sheet1' in wb.sheetnames:
        sheet1 = wb['sheet1']
    else:
        print(f"Neither 'Sheet1' nor 'sheet1' found in {file_path}")
        return

    # ---------------------------
    # Step 1: 建立新的工作表 sheet2，複製 sheet1 資料，並將第4欄(UTC時間)轉為台灣時間
    if 'sheet2' in wb.sheetnames:
        sheet2 = wb['sheet2']
    else:
        sheet2 = wb.create_sheet('sheet2')
    # 複製 sheet1 所有列（包含標題列）
    for row in sheet1.iter_rows(values_only=True):
        sheet2.append(row)
    # 將 sheet2 的第4欄(UTC)轉換為台灣時間 (從第二列開始)
    for row in sheet2.iter_rows(min_row=2, min_col=4, max_col=4):
        cell = row[0]
        if isinstance(cell.value, str):
            try:
                cell.value = utc_to_taiwan(cell.value)
            except ValueError:
                pass

    # ---------------------------
    # Step 2: 根據 sheet2 第一欄(MAC address)建立工作表，
    #         將相同 MAC 的資料複製至同一個新工作表，工作表名稱即為 MAC address
    mac_dict = {}
    for row in sheet2.iter_rows(min_row=2, max_col=4, values_only=True):
        mac = row[0]
        if mac not in mac_dict:
            mac_dict[mac] = []
        mac_dict[mac].append(row)
    # 建立 MAC 工作表（加入至 sheet2 後面）
    for mac, rows in mac_dict.items():
        ws_mac = wb.create_sheet(title=str(mac))
        # 複製標題列 (來自 sheet2)
        for col in range(1, 5):
            cell = ws_mac.cell(row=1, column=col, value=sheet2.cell(row=1, column=col).value)
            cell.font = font
            cell.alignment = alignment
            cell.border = border
        # 寫入資料列
        for data in rows:
            ws_mac.append(data)

    # ---------------------------
    # Step 3 ~ 10: 對除 sheet1、sheet2 之外的其他工作表處理
    # 排除的工作表名稱 (原 sheet1, sheet2 及 summary)
    exclude_ws = {sheet1.title, sheet2.title, 'summary'}
    # 儲存統計數據：key = 工作表名稱 (MAC)，value = (最高斷線秒數, 斷線>=60s 次數, 斷線<60s 次數, 總斷線次數, 總連線次數)
    stats = {}

    for ws in wb.worksheets:
        if ws.title in exclude_ws:
            continue

        # Step 3: 計算時間差：以第4欄時間計算與前一筆的時間差 (秒)
        # 結果放在第5欄，標題設為 'Time difference'
        ws.cell(row=1, column=5, value='Time difference')
        previous_time = None
        for row in range(2, ws.max_row + 1):
            time_val = ws.cell(row=row, column=4).value
            current_time = None
            if isinstance(time_val, str):
                try:
                    current_time = datetime.strptime(time_val, '%Y-%m-%d %H:%M:%S')
                except ValueError:
                    try:
                        current_time = datetime.strptime(time_val, '%Y-%m-%d %H:%M:%S.%f')
                    except ValueError:
                        current_time = None
            elif isinstance(time_val, datetime):
                current_time = time_val

            if current_time:
                if previous_time:
                    diff_sec = (current_time - previous_time).total_seconds()
                    diff_cell = ws.cell(row=row, column=5, value=diff_sec)
                    diff_cell.number_format = '0.00'
                previous_time = current_time

        # Step 4: 檢查第3欄是否含有 "offline"，計算總斷線次數
        offline_count = 0
        for row in range(2, ws.max_row + 1):
            if 'offline' in str(ws.cell(row=row, column=3).value).lower():
                offline_count += 1

        # 新增 Step X: 計算總連線次數，統計第3欄有出現 "connected" 的列數
        connected_count = 0
        for row in range(2, ws.max_row + 1):
            if 'connected' in str(ws.cell(row=row, column=3).value).lower():
                connected_count += 1

        # Step 6: 新增第6欄，並在 F1 儲存格設為 "斷線秒數"
        ws.cell(row=1, column=6, value='斷線秒數')
        # Step 7: 檢查每一列，若上一列第3欄包含 offline，則將本列第5欄(時間差)的數值複製到本列第6欄
        for row in range(2, ws.max_row + 1):
            prev_col3 = str(ws.cell(row=row-1, column=3).value).lower() if row-1 >= 2 else ''
            if 'offline' in prev_col3:
                ws.cell(row=row, column=6, value=ws.cell(row=row, column=5).value)
        # Step 8 ~ 10: 統計第6欄數值：
        max_disconnect = 0
        disconnect_ge60 = 0
        disconnect_lt60 = 0
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=6).value
            if isinstance(val, (int, float)):
                if val > max_disconnect:
                    max_disconnect = val
                if val >= 60:
                    disconnect_ge60 += 1
                elif val < 60:
                    disconnect_lt60 += 1
        # 記錄統計資料
        stats[ws.title] = (max_disconnect, disconnect_ge60, disconnect_lt60, offline_count, connected_count)

    # ---------------------------
    # Step 11: 建立 summary 工作表 (放在最前面)
    if 'summary' not in wb.sheetnames:
        summary_ws = wb.create_sheet('summary', 0)
    else:
        summary_ws = wb['summary']
    summary_ws.cell(row=1, column=1, value='MAC')
    summary_ws.cell(row=1, column=2, value='最高斷線秒數')
    summary_ws.cell(row=1, column=3, value='斷線>=60s次數')
    summary_ws.cell(row=1, column=4, value='斷線<60s次數')
    summary_ws.cell(row=1, column=5, value='總斷線次數')
    summary_ws.cell(row=1, column=6, value='總連線次數')  # 新增 F1 標題
    row_idx = 2
    for ws_name, (max_disc, ge60, lt60, total_offline, total_connected) in stats.items():
        summary_ws.cell(row=row_idx, column=1, value=ws_name)
        summary_ws.cell(row=row_idx, column=2, value=max_disc)
        summary_ws.cell(row=row_idx, column=3, value=ge60)
        summary_ws.cell(row=row_idx, column=4, value=lt60)
        summary_ws.cell(row=row_idx, column=5, value=total_offline)
        summary_ws.cell(row=row_idx, column=6, value=total_connected)
        row_idx += 1

    # ---------------------------
    # Step 12 & 13: 將 sheet1 改名為 Log(UTC)；將 sheet2 改名為 Log(TW)
    sheet1.title = 'Log(UTC)'
    sheet2.title = 'Log(TW)'

    # ---------------------------
    # Step 14: 為所有工作表調整字體(新細明體, size 12)、置中、加邊框，並自動調整欄寬及列高
    for ws in wb.worksheets:
        apply_formatting(ws)

    # ---------------------------
    # Step 15: 以原檔名加上 _final 儲存檔案
    new_file_path = file_path.replace('.xlsx', '_final.xlsx')
    wb.save(new_file_path)
    print(f"Processed file saved as {new_file_path}")

# 遍歷資料夾中所有 Excel 檔案
for file_name in os.listdir(folder_path):
    if file_name.endswith('.xlsx'):
        file_path = os.path.join(folder_path, file_name)
        process_excel(file_path)





# --------------------------------------------------------------------------------------------------------------------------------------------

# 程式碼說明
# Step 1 & 2：

# 讀取原始工作表（Sheet1 或 sheet1），複製至新工作表 sheet2 並將第4欄 UTC 時間轉換為台灣時間。
# 根據 sheet2 的 MAC address（第一欄）建立新的工作表，工作表名稱即為該 MAC。
# Step 3 ~ 10：

# 對除 sheet1、sheet2 之外的其他工作表（即 MAC 工作表）進行處理：
# 第3步計算每一筆資料的時間差（第4欄與前一列的時間差，結果放在第5欄）；
# 第4步統計「offline」出現的次數；
# 新增計算「總連線次數」：依據第3欄出現 "Connected" 的筆數；
# 第6步、新增第6欄 (標題「斷線秒數」)，接著第7步：若前一列第3欄含有 offline，則將本列第5欄數值複製到第6欄；
# 第8~10步則統計第6欄的最大值、斷線>=60 秒筆數及斷線<60 秒筆數。
# Step 11：

# 建立 summary 工作表，並將各 MAC 工作表的統計數據依序填入，另外 F 欄 (F1) 命名為「總連線次數」，並將各工作表中「Connected」的出現次數填入。
# Step 12 & 13：

# 將原始工作表重新命名為 Log(UTC) 與 Log(TW)。
# Step 14 & 15：

# 為所有工作表設定字型、置中、邊框及自動調整欄寬、列高，並以原檔名加上 _final 輸出檔案。
# 請將此程式碼存檔後執行，若有任何疑問或需要調整之處，再進一步告知。
