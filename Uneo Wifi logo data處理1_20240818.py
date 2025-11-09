

import openpyxl
import os

def process_excel_file(file_path):
    # 打開 Excel 檔案
    wb = openpyxl.load_workbook(file_path)
    
    # 獲取sheet1
    sheet1 = wb['sheet1']
    
    # 創建新工作表sheet2
    sheet2 = wb.create_sheet('sheet2')

    # 複製sheet1的資料到sheet2
    for row in sheet1.iter_rows(values_only=True):
        sheet2.append(row)
    
    # 將偶數列的第11至第15欄位的資料移到其前一列奇數列的第73至第77欄位
    for row_num in range(2, sheet2.max_row + 1, 2):  # 偶數列
        for col_num in range(11, 16):  # 第11至15欄
            # 將資料移到前一列奇數列的第73至77欄
            sheet2.cell(row=row_num-1, column=col_num+62).value = sheet2.cell(row=row_num, column=col_num).value
    
    # 刪除所有偶數列
    for row_num in range(sheet2.max_row, 1, -1):  # 從後往前刪
        if row_num % 2 == 0:
            sheet2.delete_rows(row_num, 1)
    
    # 保存Excel檔案，保持原始檔名
    wb.save(file_path)

# 處理多個 Excel 檔案的函式
def process_multiple_excels(folder_path):
    for filename in os.listdir(folder_path):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(folder_path, filename)
            process_excel_file(file_path)
            print(f"處理完成: {filename}")

# 使用範例
folder_path = 'C:\\_python 2024\\2024final' # Excel 檔案的資料夾路徑
process_multiple_excels(folder_path)



# # 多個excel檔, 其每個excel檔的sheet1裡的儲存格有資料, 總共72個欄位
# 1.	產出一個 新的工作表sheet2,然後把sheet1的資料複製過去
# 2.	把sheet2偶數列第11至第15個欄位的資料移到其前一列奇數列的欄位後面, 即分別移到奇數列的第73,74,75,76,77欄
# 3.	前述動作作完後,把sheet2偶數列全數刪除
# 4.	前述動作完成後, 輸出檔案並保留原始檔名
# 用Python語法


# 程式說明：
# process_excel_file(file_path) 函式：

# 開啟Excel檔案並複製Sheet1的資料到Sheet2。
# 偶數列的第11至第15欄資料移動到前一列奇數列的第73至77欄。
# 刪除所有偶數列。
# process_multiple_excels(folder_path) 函式：

# 這個函式會處理資料夾內的所有Excel檔案，並對每個檔案進行處理。
# 請將資料夾路徑替換為包含Excel檔案的路徑，並執行程式。
