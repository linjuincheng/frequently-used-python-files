# import os
# import glob
# import pandas as pd
# from openpyxl import load_workbook
# from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# # 定義樣式
# font = Font(name='新細明體', size=12)
# alignment = Alignment(horizontal='center', vertical='center')
# border = Border(left=Side(border_style='thin'), 
#                 right=Side(border_style='thin'),
#                 top=Side(border_style='thin'),
#                 bottom=Side(border_style='thin'))
# header_fill = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')  # 白色, 背景1, 較深15%

# # 設定需要處理的目錄
# input_directory = 'C:\\_python 2024\\2024final'

# # 第一部分：處理重複和缺漏的筆數
# def process_duplicates_and_missing(file_path):
#     df = pd.read_excel(file_path)
#     df['Time'] = pd.to_datetime(df.iloc[:, 4], format='%m/%d/%Y %H:%M:%S')
#     total_records = len(df)
#     duplicates = df[df.duplicated(subset=['Time'], keep=False)]
#     duplicate_count = len(duplicates)
#     df = df.sort_values(by='Time').reset_index(drop=True)
#     time_diffs = df['Time'].diff().dt.total_seconds().fillna(1)
#     missing_times_list = []
#     for i in range(1, len(df)):
#         if time_diffs[i] > 1:
#             start_time = df['Time'][i-1]
#             end_time = df['Time'][i]
#             missing_times_list.append([start_time, end_time])
#     missing_times = pd.DataFrame(missing_times_list, columns=['Start Time', 'End Time'])
#     missing_count = len(missing_times)
#     print(f"檔案: {os.path.basename(file_path)}")
#     print(f"總共有 {total_records} 筆資料")
#     print(f"時間重複的筆數: {duplicate_count} 筆")
#     print(f"時間有缺漏的筆數: {missing_count} 筆")
#     with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
#         duplicates.to_excel(writer, sheet_name='重複筆數', index=False)
#         missing_times.to_excel(writer, sheet_name='缺漏筆數', index=False)

# # 第二部分：計算缺漏時間段的秒數並設置樣式和格式
# def process_time_differences_and_style(file_path):
#     wb = load_workbook(file_path)
#     if '缺漏筆數' in wb.sheetnames:
#         ws = wb['缺漏筆數']
#         df = pd.DataFrame(ws.values)
#         df.columns = df.iloc[0]
#         df = df[1:]
#         df['Start Time'] = pd.to_datetime(df['Start Time'])
#         df['End Time'] = pd.to_datetime(df['End Time'])
#         df['time difference'] = (df['End Time'] - df['Start Time']).dt.total_seconds()
#         with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
#             df.to_excel(writer, sheet_name='缺漏筆數', index=False)
    
#     wb = load_workbook(file_path)
#     for sheet in wb.worksheets:
#         for row in sheet.iter_rows():
#             for cell in row:
#                 cell.font = font
#                 cell.alignment = alignment
#                 cell.border = border
#                 if cell.value:
#                     cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
#         first_row = sheet[1]
#         for cell in first_row:
#             if cell.value is not None:
#                 cell.fill = header_fill
#         for col in sheet.columns:
#             max_length = 0
#             column = col[0].column_letter
#             for cell in col:
#                 if cell.value:
#                     cell_length = len(str(cell.value))
#                     if cell_length > max_length:
#                         max_length = cell_length
#             adjusted_width = max_length + 2
#             sheet.column_dimensions[column].width = adjusted_width
#         for row in sheet.iter_rows():
#             max_height = 15
#             for cell in row:
#                 if cell.value:
#                     cell_length = len(str(cell.value))
#                     num_lines = (cell_length // 50) + 1
#                     height = num_lines * 15
#                     if height > max_height:
#                         max_height = height
#             sheet.row_dimensions[row[0].row].height = max_height
#     wb.save(file_path)

# # 遍歷目錄中的所有 Excel 文件並處理
# for filename in os.listdir(input_directory):
#     if filename.endswith('.xlsx'):
#         file_path = os.path.join(input_directory, filename)
#         process_duplicates_and_missing(file_path)
#         process_time_differences_and_style(file_path)



import os
import glob
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# 定義樣式
font = Font(name='新細明體', size=12)
alignment = Alignment(horizontal='center', vertical='center')
border = Border(left=Side(border_style='thin'), 
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin'))
header_fill = PatternFill(start_color='D0D0D0', end_color='D0D0D0', fill_type='solid') 

# 設定需要處理的目錄
input_directory = 'C:\\_python 2024\\2024final'

# 第一部分：處理重複和缺漏的筆數
def process_duplicates_and_missing(file_path):
    df = pd.read_excel(file_path)
    df['Time'] = pd.to_datetime(df.iloc[:, 4], format='%m/%d/%Y %H:%M:%S')
    total_records = len(df)
    duplicates = df[df.duplicated(subset=['Time'], keep=False)]
    duplicate_count = len(duplicates)
    df = df.sort_values(by='Time').reset_index(drop=True)
    time_diffs = df['Time'].diff().dt.total_seconds().fillna(1)
    missing_times_list = []
    for i in range(1, len(df)):
        if time_diffs[i] > 1:
            start_time = df['Time'][i-1]
            end_time = df['Time'][i]
            missing_times_list.append([start_time, end_time])
    missing_times = pd.DataFrame(missing_times_list, columns=['Start Time', 'End Time'])
    missing_count = len(missing_times)
    print(f"檔案: {os.path.basename(file_path)}")
    print(f"總共有 {total_records} 筆資料")
    print(f"時間重複的筆數: {duplicate_count} 筆")
    print(f"時間有缺漏的筆數: {missing_count} 筆")
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        duplicates.to_excel(writer, sheet_name='重複筆數', index=False)
        missing_times.to_excel(writer, sheet_name='缺漏筆數', index=False)

# 第二部分：計算缺漏時間段的秒數並設置樣式和格式
def process_time_differences_and_style(file_path):
    wb = load_workbook(file_path)
    if '缺漏筆數' in wb.sheetnames:
        ws = wb['缺漏筆數']
        df = pd.DataFrame(ws.values)
        df.columns = df.iloc[0]
        df = df[1:]
        df['Start Time'] = pd.to_datetime(df['Start Time'])
        df['End Time'] = pd.to_datetime(df['End Time'])
        df['time difference'] = (df['End Time'] - df['Start Time']).dt.total_seconds()
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name='缺漏筆數', index=False)
    
    wb = load_workbook(file_path)
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = font
                cell.alignment = alignment
                cell.border = border
                if cell.value:
                    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        first_row = sheet[1]
        for cell in first_row:
            if cell.value is not None:
                cell.fill = header_fill
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            adjusted_width = max_length + 2
            sheet.column_dimensions[column].width = adjusted_width
        for row in sheet.iter_rows():
            max_height = 15
            for cell in row:
                if cell.value:
                    cell_length = len(str(cell.value))
                    num_lines = (cell_length // 50) + 1
                    height = num_lines * 15
                    if height > max_height:
                        max_height = height
            sheet.row_dimensions[row[0].row].height = max_height
    wb.save(file_path)

# 遍歷目錄中的所有 Excel 文件並處理
for filename in os.listdir(input_directory):
    if filename.endswith('.xlsx'):
        file_path = os.path.join(input_directory, filename)
        process_duplicates_and_missing(file_path)
        process_time_differences_and_style(file_path)


