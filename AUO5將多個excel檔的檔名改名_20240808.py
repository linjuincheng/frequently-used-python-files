# 檔名依sheet1工作表的第一列的MAC及Time欄位裡的日期作相對應的變化, 檔名格式AUO_MAC_TW_Time 
# e.g.如果MAC欄位是的第二列是D83ADD861851, 而Time(format='%m/%d/%Y %H:%M:%S')欄位的第二列是08/06/2024 10:03:02 
# 則excel檔名輸出檔名改為AUO_D83ADD861851_TW_20240806並取代原檔案

# 說明：
# 尋找 MAC 和 Time 欄位：

# 程式會檢查 sheet1 的第一列，以找到 MAC 和 Time 欄位的列號。
# 提取 MAC 和 Time 值：

# 假設資料在第二列（即第二行）。提取 MAC 和 Time 欄位的值。
# 轉換 Time 格式：

# 將 Time 的值轉換為 YYYYMMDD 格式的字串。
# 生成新檔名：

# 使用 MAC 和日期生成新檔案名稱，格式為 AUO_MAC_TW_YYYYMMDD.xlsx。
# 重命名檔案：

# 使用 os.rename 將檔案重命名為新的檔案名稱，並取代原始檔案。
# 注意事項：
# 這段程式碼會在目錄中的所有 .xlsx 檔案上運行。如果檔案名稱重複，則會覆蓋檔案，請小心使用。
# 程式假設資料在第二列，如果資料位置不同，請調整相應的 row 參數。

import os
import pandas as pd
from openpyxl import load_workbook

# 設定資料夾路徑
folder_path = 'C:\\_python 2024\\2024final'

# 遍歷資料夾中的每個 Excel 檔案
for filename in os.listdir(folder_path):
    if filename.endswith('.xlsx'):
        file_path = os.path.join(folder_path, filename)

        # 讀取 Excel 檔案中的 sheet1
        workbook = load_workbook(file_path, data_only=True)
        if 'sheet1' in workbook.sheetnames:
            sheet = workbook['sheet1']

            # 查找 MAC 和 Time 欄位的列號
            mac_col = None
            time_col = None
            for col in range(1, sheet.max_column + 1):
                header_value = sheet.cell(row=1, column=col).value
                if header_value == 'MAC':
                    mac_col = col
                elif header_value == 'Time':
                    time_col = col

            if mac_col and time_col:
                # 假設數據在第二列
                mac = sheet.cell(row=2, column=mac_col).value
                time_value = sheet.cell(row=2, column=time_col).value

                if mac and time_value:
                    # 轉換時間格式到 YYYYMMDD
                    date_str = pd.to_datetime(time_value).strftime('%Y%m%d')

                    # 生成新檔名
                    new_filename = f"AUO_{mac}_TW_{date_str}.xlsx"
                    new_file_path = os.path.join(folder_path, new_filename)

                    # 重命名檔案，取代原檔案
                    os.rename(file_path, new_file_path)
                    print(f"已將 {filename} 重命名為 {new_filename}")

print("所有文件處理完畢")

