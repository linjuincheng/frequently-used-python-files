

import openpyxl
import os

def process_excel_file(file_path):
    # 打開 Excel 檔案
    wb = openpyxl.load_workbook(file_path)
    
    # 獲取 sheet1
    sheet1 = wb['sheet1']
    
    # 創建新工作表 sheet2 並複製 sheet1 資料
    sheet2 = wb.create_sheet('sheet2')
    for row in sheet1.iter_rows(values_only=True):
        sheet2.append(row)
    
    # 將偶數列的第11至第15欄資料移到奇數列的第73至第77欄
    for row_num in range(2, sheet2.max_row + 1, 2):  # 偶數列
        for col_num in range(11, 16):  # 第11至15欄
            sheet2.cell(row=row_num-1, column=col_num+62).value = sheet2.cell(row=row_num, column=col_num).value
    
    # 刪除偶數列
    for row_num in range(sheet2.max_row, 1, -1):  # 從後往前刪除
        if row_num % 2 == 0:
            sheet2.delete_rows(row_num)
    
    # 創建 sheet3 並複製 sheet2 資料
    sheet3 = wb.create_sheet('sheet3')
    for row in sheet2.iter_rows(values_only=True):
        sheet3.append(row)
    
    # 在 sheet3 最上方插入空白列
    sheet3.insert_rows(1)
    
    # 將 sheet3 第76欄的數字轉換為16進位
    for row in range(2, sheet3.max_row + 1):
        cell_value = sheet3.cell(row=row, column=76).value
        if isinstance(cell_value, (int, float)):  # 確認是數字
            sheet3.cell(row=row, column=76).value = hex(int(cell_value))[2:].upper()
    
    # 創建 sheet4 並複製 sheet3 資料
    sheet4 = wb.create_sheet('sheet4')
    for row in sheet3.iter_rows(values_only=True):
        sheet4.append(row)
    
    # 在 sheet4 最上方插入空白列
    sheet4.insert_rows(1)
    
    # 如果第76欄是 'D' 或 'E'，進行替換
    for row in range(2, sheet4.max_row + 1):
        cell_value = sheet4.cell(row=row, column=76).value
        if cell_value == 'D':
            sheet4.cell(row=row, column=76).value = 'Sensor error'
        elif cell_value == 'E':
            sheet4.cell(row=row, column=76).value = 'Wifi error'
    
    # 保存 Excel 檔案並保持原始檔名
    wb.save(file_path)

# 處理多個 Excel 檔案的函式
def process_multiple_excels(folder_path):
    for filename in os.listdir(folder_path):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(folder_path, filename)
            process_excel_file(file_path)
            print(f"處理完成: {filename}")

# 使用範例
folder_path = r"C:\_python 2024\2024final"  # Excel 檔案的資料夾路徑
process_multiple_excels(folder_path)


# 多個excel檔(其預設路徑放在C:\_python 2024\2024final), 其每個excel檔的sheet1裡的儲存格有資料, 總共72個欄位
# 1.	產出一個 新的工作表sheet2,然後把sheet1的資料複製過去
# 2.	把sheet2偶數列第11至第15個欄位的資料移到其前一列奇數列的欄位後面, 即分別移到奇數列的第73,74,75,76,77欄
# 3.	前述動作作完後,把sheet2偶數列全數刪除
# 4.	前述動作作完後, 產出一個 新的工作表sheet3,然後把sheet2的資料複製過去,並在sheet3最上方新增一列空白列並把sheet3第76欄的數字換算成16 進位的數字(或文字)
# 5.	前述動作作完後, 產出一個 新的工作表sheet4,然後把sheet3的資料複製過去,並在sheet4最上方新增一列空白列,如果第76欄的儲存格內容是D則把儲存格內容換為Sensor error, 如是E,則則把儲存格內容換為Wifi error
# 6.	前述動作完成後, 輸出檔案並保留原始檔名
# 用Python語法


# 程式步驟解釋：
# 複製 Sheet1 到 Sheet2：將原始工作表的資料完整複製到新工作表 Sheet2 中。
# 資料搬移：將 Sheet2 偶數列第11至第15欄的資料移動到前一行奇數列的第73至第77欄。
# 刪除偶數列：移動資料後，刪除 Sheet2 中所有偶數列。
# 複製到 Sheet3 並新增空白列：將 Sheet2 的資料複製到 Sheet3，並在最上方新增一行空白列，同時將 Sheet3 中的第76欄數字轉換為16進位。
# 複製到 Sheet4 並替換特定值：將 Sheet3 的資料複製到 Sheet4，並在 Sheet4 中替換第76欄中符合條件的資料（D 替換為 "Sensor error"，E 替換為 "Wifi error"）。
# 保存檔案：最終將修改後的 Excel 檔案保存，並保持原始檔名不變。
# 這段程式會處理資料夾中的所有 .xlsx 檔案並應用相同的處理邏輯。如果有其他問題或需求可以再調整。