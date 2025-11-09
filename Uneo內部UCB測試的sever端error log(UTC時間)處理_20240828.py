

import os
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

# 設定資料夾路徑
folder_path = r'C:\_python 2024\2024final'

# 定義字體、對齊和邊框樣式
font = Font(name='新細明體', size=12)
alignment = Alignment(horizontal='center', vertical='center')
border = Border(left=Side(border_style='thin'), 
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin'))

def utc_to_taiwan(utc_str):
    """Convert UTC time to Taiwan time."""
    try:
        utc_time = datetime.strptime(utc_str, '%Y-%m-%d %H:%M:%S.%f')
    except ValueError:
        try:
            utc_time = datetime.strptime(utc_str, '%Y-%m-%d %H:%M:%S')
        except ValueError:
            raise ValueError(f"Unsupported time format: {utc_str}")
    
    taiwan_time = utc_time + timedelta(hours=8)
    return taiwan_time.strftime('%Y-%m-%d %H:%M:%S')

def process_excel(file_path):
    wb = load_workbook(file_path)

    # 確認是否存在 'Sheet1' 或 'sheet1'
    if 'Sheet1' in wb.sheetnames:
        sheet1 = wb['Sheet1']
    elif 'sheet1' in wb.sheetnames:
        sheet1 = wb['sheet1']
    else:
        print(f"Neither 'Sheet1' nor 'sheet1' found in {file_path}")
        return

    # Step 1: 產出 sheet2 並轉換 UTC 時間
    if 'sheet2' in wb.sheetnames:
        sheet2 = wb['sheet2']
    else:
        sheet2 = wb.create_sheet('sheet2')

    # 複製資料到 sheet2
    for row in sheet1.iter_rows(values_only=True):
        sheet2.append(row)

    # 轉換 UTC 時間為台灣時間
    for row in sheet2.iter_rows(min_row=2, max_col=4):
        utc_time_cell = row[3].value
        if isinstance(utc_time_cell, str):
            try:
                taiwan_time = utc_to_taiwan(utc_time_cell)
                row[3].value = taiwan_time
            except ValueError:
                row[3].value = utc_time_cell  # Keep original if format is unsupported

    # Step 2: 根據 MAC 地址創建新的工作表
    mac_dict = {}
    for row in sheet2.iter_rows(min_row=2, max_col=4, values_only=True):
        mac = row[0]
        if mac not in mac_dict:
            mac_dict[mac] = []
        mac_dict[mac].append(row)

    for mac, rows in mac_dict.items():
        ws = wb.create_sheet(mac)
        for col_index in range(1, 5):
            cell = ws.cell(row=1, column=col_index, value=sheet2.cell(row=1, column=col_index).value)
            cell.font = font
            cell.alignment = alignment
            cell.border = border
        for row in rows:
            ws.append(row)

    # Step 3 & 4: 計算時間差和檢查 "offline"
    offline_counts = {}
    for ws in wb.worksheets:
        if ws.title in ['Sheet1', 'sheet2', 'summary']:
            continue
        
        previous_time = None
        offline_count = 0

        # Step 3: 計算時間差
        ws.cell(row=1, column=5, value='Time difference')
        for row in range(2, ws.max_row + 1):
            current_time_str = ws.cell(row=row, column=4).value
            if isinstance(current_time_str, str):
                try:
                    current_time = datetime.strptime(current_time_str, '%Y-%m-%d %H:%M:%S')
                except ValueError:
                    try:
                        current_time = datetime.strptime(current_time_str, '%Y-%m-%d %H:%M:%S.%f')
                    except ValueError:
                        continue  # Skip rows with invalid date formats

                if previous_time:
                    time_diff = (current_time - previous_time).total_seconds()
                    ws.cell(row=row, column=5, value=time_diff).number_format = '0.00'
                
                previous_time = current_time

            # Step 4: 檢查 "offline"
            if 'offline' in str(ws.cell(row=row, column=3).value).lower():
                offline_count += 1
        
        offline_counts[ws.title] = offline_count

    # Step 5: 填寫 Summary 工作表
    if 'summary' not in wb.sheetnames:
        summary_ws = wb.create_sheet('summary', 0)
    else:
        summary_ws = wb['summary']
    
    summary_ws.cell(row=1, column=1, value='File Name')
    summary_ws.cell(row=1, column=2, value='Offline Count')
    
    row_index = 2
    for ws_title, offline_count in offline_counts.items():
        summary_ws.cell(row=row_index, column=1, value=ws_title)
        summary_ws.cell(row=row_index, column=2, value=offline_count)
        row_index += 1

    # Step 6: 設置所有工作表的字體、對齊和邊框，並自動調整欄位寬度和列高
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.font = font
                cell.alignment = alignment
                cell.border = border
        
        # 自動調整列高和欄位寬度
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if isinstance(cell.value, str):
                    if len(cell.value) > max_length:
                        max_length = len(cell.value)
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        for row in ws.iter_rows():
            max_height = 0
            for cell in row:
                if cell.value:
                    height = len(str(cell.value).split('\n'))
                    if height > max_height:
                        max_height = height
            ws.row_dimensions[row[0].row].height = (max_height + 2) * 15

    # 儲存修改後的文件
    new_file_path = file_path.replace('.xlsx', '_final.xlsx')
    wb.save(new_file_path)
    print(f"Processed file saved as {new_file_path}")

# 遍歷資料夾中的所有 Excel 檔案
for file_name in os.listdir(folder_path):
    if file_name.endswith('.xlsx'):
        file_path = os.path.join(folder_path, file_name)
        process_excel(file_path)
        
        
        
        
        
        
#     多個excel檔(其預設路徑放在C:\_python 2024\2024final), 其每個excel檔的sheet1或者Sheet1工作表裡的有資料, 總共4個欄位, 第1個欄位是MAC address, 第4欄位為UTC時間 
# 1. 產出新的工作表取名sheet2, 將shee1的資料複製到sheet2並直接把sheet2第4欄位的UTC時間改為台灣時間
# 2. 將sheet2裡第1個欄位裡有同樣MAC address的儲存格內容篩選出來, 同樣MAC address 產出一個新的工作表,工作表接續在sheet2之後, 並把工作表的名稱以MAC address的名稱命名(e.g. 80C9555CC3CC)
# 2. 忽略sheet1及sheet2, 計算sheet1及sheet2之外的其他每個工作表的第4個欄位(時間欄位)的每一列(筆)與其前1列(筆)資料的時間差距(sheet1及sheet2及summary這3個工作表都不用計算時間差距)       
# 3. 忽略sheet1及sheet2, 檢查sheet1及sheet2之外的其他每個工作表的第3個欄位的每一列資料, 看是否資料裡帶有offline
# 4. 計算有每個工作表有多少列的資料帶有offline,則為斷線次數
# 5. 在excel檔所有工作前最前面產出名為summary的工作表, 把各工作表的檔名放在summary工作表的第1欄, 計算出的斷線次數放在summary 工作表的第2欄
# 6. 調整每個儲存格的字體(新細明體,字形size12,儲存格置中並加邊框, 以及對工作表內容自動調整欄寬,自動調整列高
# 7. 在原檔名後加_final後輸出執行程式後的檔案

# 用Python語法


# 說明：
# utc_to_taiwan 函數：

# 處理 UTC 時間轉換，增加了對毫秒的處理。
# process_excel 函數：

# Step 1: 產生 sheet2 並將 Sheet1 的資料複製過去，並轉換時間為台灣時間。
# Step 2: 根據 sheet2 的 MAC 地址創建新的工作表。
# Step 3 & 4: 計算時間差，檢查 offline 並計算斷線次數。
# Step 5: 在 summary 工作表中記錄每個工作表的檔名和斷線次數。
# Step 6: 調整儲存格的字體、對齊、邊框，並自動調整欄寬和列高。
# 遍歷資料夾中的所有 Excel 檔案：

# 讀取每個 Excel 檔案並處理。
# 將此程式碼保存為 .py 檔案並執行，它會自動處理指定資料








