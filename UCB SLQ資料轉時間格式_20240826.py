# -*- coding: utf-8 -*-
"""
Created on Mon Aug 26 17:39:25 2024

@author: user
"""

import os
from openpyxl import load_workbook
import pandas as pd

# 設定資料夾路徑
folder_path = r'C:\_python 2024\2024final'

# 定義函數以更新時間格式
def update_time_format(file_path):
    # 加載 Excel 檔案
    workbook = load_workbook(file_path)
    sheet = workbook['sheet1']

    # 假設時間欄位在第4欄，並從第二行開始
    for row in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=4).value
        if cell_value:
            try:
                # 將時間格式改為 '%m/%d/%Y %H:%M:%S'
                new_time = pd.to_datetime(cell_value).strftime('%m/%d/%Y %H:%M:%S')
                sheet.cell(row=row, column=4).value = new_time
            except Exception as e:
                print(f"無法解析時間值 {cell_value}，位於檔案 {file_path} 的第 {row} 行。錯誤: {e}")

    # 儲存檔案並保持原檔名
    workbook.save(file_path)
    print(f"已處理完成並儲存：{file_path}")

# 遍歷資料夾中的每個 Excel 檔案
for filename in os.listdir(folder_path):
    if filename.endswith('.xlsx'):
        file_path = os.path.join(folder_path, filename)
        update_time_format(file_path)

print("所有文件處理完畢")


# 多個excel檔(其預設路徑放在C:\_python 2024\2024final),
# 在sheet1工作作表的第4欄是時間欄位, 需把時間欄位的時間格式改為'%m/%d/%Y %H:%M:%S'格式
# 輸出的excel檔名維持原檔名
# 用Python語法


# 說明：
# 路徑設置：程式會遍歷 C:\_python 2024\2024final 資料夾中的所有 .xlsx 檔案。
# 時間格式更新：針對每個檔案中的 sheet1，我們將第 4 欄的時間數據轉換為 '%m/%d/%Y %H:%M:%S' 格式。
# 保持原檔名：修改後的檔案會以原檔名儲存。
# 注意事項：
# 假設所有 Excel 檔案都有名為 sheet1 的工作表，且第 4 欄是時間數據。
# 如果遇到無法解析的時間值，會跳過該單元格並顯示錯誤訊息。
# 這個程式會自動遍歷指定資料夾中的所有 .xlsx 檔案，並更新時間格式。
