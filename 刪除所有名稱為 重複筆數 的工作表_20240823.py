import os
from openpyxl import load_workbook

# 設定資料夾路徑
folder_path = r'C:\_python 2024\2024final'

# 取得資料夾中的所有 Excel 檔案
excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]

# 遍歷每個 Excel 檔案
for file in excel_files:
    file_path = os.path.join(folder_path, file)
    
    # 讀取 Excel 檔案
    workbook = load_workbook(file_path)
    
    # 檢查並刪除名稱為 '重複筆數' 的工作表
    if '重複筆數' in workbook.sheetnames:
        del workbook['重複筆數']
        # 儲存修改過的 Excel 檔案
        workbook.save(file_path)
        
print("工作表刪除完成！")



# 多個excel檔(其預設路徑放在C:\_python 2024\2024final), 刪除excel檔裡名稱為重複筆數的工作表
# 用Python語法


# 程式碼解釋：
# os.listdir(folder_path)：列出指定資料夾中的所有檔案。
# load_workbook(file_path)：使用 openpyxl 讀取 Excel 檔案。
# if '重複筆數' in workbook.sheetnames：檢查工作簿中是否存在名為 重複筆數 的工作表。
# del workbook['重複筆數']：刪除名為 重複筆數 的工作表。
# workbook.save(file_path)：儲存修改過的 Excel 檔案。
# 這段程式碼將會自動遍歷指定資料夾中的所有 Excel 檔案，並刪除每個檔案中名為 重複筆數 的工作表。