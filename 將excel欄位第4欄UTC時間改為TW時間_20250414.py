import os
from openpyxl import load_workbook
from datetime import datetime, timedelta

# 資料夾路徑
folder_path = r"C:\_python 2024\2024final"
# 定義字串格式（若時間以字串儲存）
time_format = "%Y-%m-%d %H:%M:%S.%f"

# 遍歷指定資料夾內所有 .xlsx 檔案
for filename in os.listdir(folder_path):
    if filename.endswith(".xlsx"):
        file_path = os.path.join(folder_path, filename)
        wb = load_workbook(file_path)
        
        # 尋找工作表名稱中符合 "sheet1" (不分大小寫)
        sheet1 = None
        for s in wb.sheetnames:
            if s.lower() == "sheet1":
                sheet1 = wb[s]
                break
        if sheet1 is None:
            print(f"檔案 {filename} 中找不到 'Sheet1'，跳過檔案。")
            continue

        # 新建工作表, 暫名 "sheet2"，並複製 sheet1 的資料
        sheet2 = wb.create_sheet(title="sheet2")
        for row in sheet1.iter_rows(values_only=True):
            sheet2.append(row)
            
        # 於 sheet2 中對第4欄(也就是 D 欄)進行時間轉換
        for row in range(1, sheet2.max_row + 1):
            cell = sheet2.cell(row=row, column=4)
            if cell.value is not None:
                # 若內容已為 datetime 物件，直接加上 8 小時
                if isinstance(cell.value, datetime):
                    cell.value = cell.value + timedelta(hours=8)
                # 若內容為字串，嘗試解析後再轉換
                elif isinstance(cell.value, str):
                    try:
                        dt_obj = datetime.strptime(cell.value, time_format)
                        cell.value = dt_obj + timedelta(hours=8)
                    except Exception as e:
                        print(f"檔案 {filename} 轉換時間錯誤，儲存格 {cell.coordinate} 內容：{cell.value}, 錯誤：{e}")
                else:
                    print(f"檔案 {filename} 儲存格 {cell.coordinate} 的資料型態非預期，請檢查資料。")
        
        # 重新命名工作表：原 sheet1 改為 "Log(UTC)"，sheet2 改為 "Log(TW)"
        sheet1.title = "Log(UTC)"
        sheet2.title = "Log(TW)"
        
        # 輸出新檔案，檔名在原檔名後加上 "_final"
        new_filename = filename.replace(".xlsx", "_final.xlsx")
        new_file_path = os.path.join(folder_path, new_filename)
        wb.save(new_file_path)
        print(f"檔案 {filename} 處理完畢，已輸出: {new_filename}")



# -------------------------------------------------------------------------------------------------------------

# 多個excel檔(其預設路徑放在C:\_python 2024\2024final), 其每個excel檔的sheet1或者Sheet1工作表裡的有資料, 總共4個欄位, 第1個欄位是MAC address, 第4欄位為UTC時間 

# 1.	產出新的工作表取名sheet2, 將shee1的資料複製到sheet2並直接把sheet2第4欄位的UTC時間改為台灣時間
# 2.	將sheet1改名為Log(UTC)
# 3.	將shee2   改名為 Log(TW_
# 4.	在原檔名後加_final後輸出執行程式後的檔案
# 以上,用Python語法


# ---------------------------------------------------------------------------------------------------------------------------

# 程式碼說明
# 讀取及尋找工作表

# 利用 os.listdir() 遍歷指定資料夾內所有 .xlsx 檔案。

# 使用 wb.sheetnames 與不區分大小寫的比對，找到名稱為 "Sheet1" 的工作表。

# 複製資料至新工作表

# 建立新的工作表(暫命名 "sheet2")，並使用 iter_rows(values_only=True) 逐列將資料複製到新工作表中。

# 時間轉換 (UTC → 台灣時間)

# 直接遍歷新工作表中第4欄的所有儲存格，透過 sheet2.cell(row=row, column=4) 取得 cell 物件。

# 若該儲存格內容為 datetime 物件則直接加 8 小時；

# 若內容為字串，則先依指定格式解析成 datetime 物件再加 8 小時。

# 若遇到其他型態，則印出提示訊息。

# 重新命名與儲存檔案

# 將原 "Sheet1" 改名為 "Log(UTC)"，新工作表改名為 "Log(TW)"。

# 儲存時於原檔名後加上 _final 後綴，產生新檔案。

# 請確保系統已安裝 openpyxl (pip install openpyxl) 並注意 Excel 檔中第4欄時間資料的格式是否與程式中設定的 time_format 相符。