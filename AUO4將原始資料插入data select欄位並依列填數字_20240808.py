

import openpyxl
import os

# 設定要處理的Excel檔案夾路徑
folder_path = 'C:\\_python 2024\\2024final'  # 替換為包含Excel檔案的資料夾路徑

# 遍歷資料夾中的每個Excel檔案
for filename in os.listdir(folder_path):
    if filename.endswith('.xlsx'):
        file_path = os.path.join(folder_path, filename)
        
        # 載入Excel檔案
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook['sheet1']  # 假設每個檔案都有一個名為Sheet1的工作表

        # 在第1個欄位前插入新的欄位
        sheet.insert_cols(1)

        # 在A1儲存格填上"Data select"
        sheet['A1'] = 'Data select'

        # 從0開始計數，並在第2個欄位有資料的儲存格前填上相對應的數字
        counter = 0
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=2).value is not None:  # 檢查第2個欄位是否有資料
                sheet.cell(row=row, column=1).value = counter  # 在第1個欄位填上數字
                counter += 1

        # 儲存修改後的Excel檔案
        workbook.save(file_path)
        print(f'{filename} 處理完成')

print("所有文件處理完畢")
