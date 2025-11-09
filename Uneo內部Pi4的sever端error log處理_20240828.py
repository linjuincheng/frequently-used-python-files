


import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

# 設定資料夾路徑
folder_path = r'C:\_python 2024\2024final'

# 取得資料夾中的所有 Excel 檔案
excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]

# 定義樣式
font = Font(name='新細明體', size=12)
alignment = Alignment(horizontal='center', vertical='center')
border = Border(left=Side(border_style='thin'), 
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin'))

# 逐個處理 Excel 檔案
for file in excel_files:
    file_path = os.path.join(folder_path, file)
    
    # 讀取 Excel 檔案的 sheet1
    df = pd.read_excel(file_path, sheet_name='sheet1')
    
    # 1. 計算第4個欄位的時間差並將結果放在第5個欄位
    df['Time'] = pd.to_datetime(df.iloc[:, 3])  # 確保第4個欄位為時間格式
    df['Time_Diff'] = df['Time'].diff().dt.total_seconds().fillna(0)  # 計算時間差並填入第5個欄位
    
    # 2. 依據 MAC address 分組並生成新的工作表
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        # 先將處理過的 sheet1 回寫
        df.to_excel(writer, sheet_name='sheet1', index=False)
        
        # 根據 MAC address 分組，並將每個組寫入新的工作表
        for mac_address, group in df.groupby(df.iloc[:, 0]):
            group.to_excel(writer, sheet_name=str(mac_address), index=False)
    
    # 3. 打開 Excel 檔案以便進行後續操作
    wb = load_workbook(file_path)
    
    # 4. 建立 summary 工作表
    if 'summary' in wb.sheetnames:
        ws_summary = wb['summary']
    else:
        ws_summary = wb.create_sheet(title='summary', index=0)
    
    # 設置 summary 標題
    ws_summary['A1'] = 'Sheet Name'
    ws_summary['B1'] = 'Offline Count'
    ws_summary['A1'].font = font
    ws_summary['B1'].font = font

    summary_row = 2  # summary 工作表的行計數起始
    
    # 遍歷所有工作表並應用樣式
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        
        # 忽略 'sheet1' 和 'summary'
        if sheet in ['sheet1', 'summary']:
            continue
        
        # 計算 offline 次數
        offline_count = 0
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):  # 從第2行開始避免標題行
            cell_value = row[0].value
            if cell_value and 'offline' in str(cell_value).lower():
                offline_count += 1
        
        # 將結果寫入 summary 工作表
        ws_summary[f'A{summary_row}'] = sheet
        ws_summary[f'B{summary_row}'] = offline_count
        ws_summary[f'A{summary_row}'].font = font
        ws_summary[f'B{summary_row}'].font = font
        summary_row += 1
        
        # 應用字體、對齊和邊框樣式
        for row in ws.iter_rows():
            for cell in row:
                cell.font = font
                cell.alignment = alignment
                cell.border = border
        
        # 自動調整欄寬
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # 取得欄位字母
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
        
        # 自動調整列高
        for row in ws.iter_rows():
            max_height = 0
            for cell in row:
                if cell.value:
                    cell_height = len(str(cell.value)) // 20 + 1
                    max_height = max(max_height, cell_height)
            ws.row_dimensions[row[0].row].height = max_height * 15

    # 儲存 Excel 檔案
    new_file_path = os.path.join(folder_path, file.replace('.xlsx', '_final.xlsx'))
    wb.save(new_file_path)
    
    
    
#     多個excel檔(其預設路徑放在C:\_python 2024\2024final), 其每個excel檔的sheet1裡的儲存格有資料, 總共4個欄位, 第1個欄位是MAC address, 第4欄位為台灣時間 
# 1.	計算shee1的第4個欄位的時間欄位的每一列(筆)與其前1列(筆)資料的時間差距並把結果列在shee1的第5個欄位
# 2.	將sheet1裡1個欄位裡同樣MAC address的篩選出來, 同樣MAC address產出一個新的工作表 , 工作表接續在sheet1之後, 並把工作表的名稱以MAC address的名稱命名(e.g. 80C9555CC3CC)
# 3.	將所有工作表的調整字體及欄位, 如下列程式碼
# font = Font(name='新細明體', size=12)
# alignment = Alignment(horizontal='center', vertical='center')
# border = Border(left=Side(border_style='thin'), 
#                 right=Side(border_style='thin'),
#                 top=Side(border_style='thin'),
#                 bottom=Side(border_style='thin'))
# 4. 忽略sheet1, 檢查sheet1之外的其它每一個工作表的第三個欄位的每一列資料, 看是否資料裡帶有offline
# 5. 計算有每個工作表有多少列的資料帶有offline,則為斷線次數
# 6. 在excel檔所有工作前最前面產出名為summary的工作表, 把各工作表的檔名放在summary工作表的第1欄, 計算出的斷線次數放在summary 工作表的第2欄
# 7. 執行上述程式碼產出各工作表後, 還是把所有工作表的字體設為新細明體', size=12, 儲存格欄位置中對齊,自動調整列高及自動調整欄高後再輸出, 輸出檔名在原檔名後加_final後輸出
# 用Python語法



# 時間差計算：

# 程式會在 sheet1 中計算第4欄位的時間差，並將結果寫入第5欄位。
# 按 MAC address 分組：

# 將 sheet1 中相同 MAC address 的資料分組，每個組生成一個新的工作表，並將工作表名稱設為該 MAC address。
# 樣式應用：

# 將所有的工作表，包括 sheet1 和新生成的工作表，字體設置為 '新細明體'，字體大小為 12，儲存格內容居中，並且為每個儲存格添加邊框。
# 離線次數計算：

# 程式會檢查每個工作表的第三欄位是否包含 offline，並計算出現的次數。
# Summary 工作表：

# 程式會在所有工作表的最前面新增一個 summary 工作表，將每個工作表的名稱和計算出的斷線次數填入 summary 工作表。
# 檔案輸出：

# 最後，程式會將處理後的 Excel 檔案以原檔名加上 _final 的格式保存。
# 此程式會遍歷資料夾中的所有 Excel 檔案，並對每個檔案進行上述處理，最終生成新的 Excel 檔案，保存在同一資料夾中。


