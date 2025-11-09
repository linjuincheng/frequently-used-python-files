import os
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

# 設定資料夾路徑
folder_path = r'C:\_python 2024\2024final'

# 定義字體、對齊和邊框樣式
font = Font(name='新細明體', size=12)
alignment = Alignment(horizontal='center', vertical='center')
border = Border(left=Side(border_style='thin'), 
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin'))

def utc_to_taiwan(utc_str):
    """Convert UTC time to Taiwan time."""
    try:
        utc_time = datetime.strptime(utc_str, '%Y-%m-%d %H:%M:%S.%f')
    except ValueError:
        try:
            utc_time = datetime.strptime(utc_str, '%Y-%m-%d %H:%M:%S')
        except ValueError:
            raise ValueError(f"Unsupported time format: {utc_str}")
    
    taiwan_time = utc_time + timedelta(hours=8)
    return taiwan_time.strftime('%Y-%m-%d %H:%M:%S')

def process_excel(file_path):
    wb = load_workbook(file_path)

    # 確認是否存在 'Sheet1' 或 'sheet1'
    sheet_name = next((s for s in ['Sheet1', 'sheet1'] if s in wb.sheetnames), None)
    if sheet_name is None:
        print(f"Neither 'Sheet1' nor 'sheet1' found in {file_path}")
        return
    
    sheet1 = wb[sheet_name]

    # Step 1: 產出 sheet2 並轉換 UTC 時間
    if 'sheet2' in wb.sheetnames:
        sheet2 = wb['sheet2']
    else:
        sheet2 = wb.create_sheet('sheet2')

    # 複製 sheet1 的資料到 sheet2
    for row in sheet1.iter_rows():
        new_row = [cell.value for cell in row]
        sheet2.append(new_row)

    # 轉換 sheet2 第4欄(UTC時間)為台灣時間 (從第二列開始)
    for row in sheet2.iter_rows(min_row=2, min_col=4, max_col=4):
        cell = row[0]
        if isinstance(cell.value, str):
            try:
                cell.value = utc_to_taiwan(cell.value)
            except ValueError:
                # 無法轉換則保持原值
                pass

    # Step 2: 根據 MAC 地址 (sheet2 第1欄) 創建新的工作表
    mac_dict = {}
    for row in sheet2.iter_rows(min_row=2, max_col=4, values_only=True):
        mac = row[0]
        if mac not in mac_dict:
            mac_dict[mac] = []
        mac_dict[mac].append(row)

    for mac, rows in mac_dict.items():
        ws = wb.create_sheet(mac)
        # 複製標題列 (取自 sheet2 第一列)
        for col_index in range(1, 5):
            cell = ws.cell(row=1, column=col_index, value=sheet2.cell(row=1, column=col_index).value)
            cell.font = font
            cell.alignment = alignment
            cell.border = border
        # 寫入該 MAC 對應的資料列
        for r in rows:
            ws.append(r)

    # Step 3 & 4: 計算時間差、檢查 "offline" 與統計 "connected"
    offline_counts = {}
    connected_counts = {}

    # 針對除 'Sheet1'/'sheet1'、'sheet2'、及 'summary' 之外的工作表
    for ws in wb.worksheets:
        if ws.title in ['Sheet1', 'sheet1', 'sheet2', 'summary']:
            continue

        previous_time = None
        offline_count = 0
        connected_count = 0
        connected_rows = 0

        # Step 3: 計算時間差，將結果放到第5欄，第一列設為標題
        ws.cell(row=1, column=5, value='Time difference')
        for row in range(2, ws.max_row + 1):
            current_time_str = ws.cell(row=row, column=4).value
            current_time = None
            if isinstance(current_time_str, str):
                try:
                    current_time = datetime.strptime(current_time_str, '%Y-%m-%d %H:%M:%S')
                except ValueError:
                    try:
                        current_time = datetime.strptime(current_time_str, '%Y-%m-%d %H:%M:%S.%f')
                    except ValueError:
                        current_time = None
            elif isinstance(current_time_str, datetime):
                current_time = current_time_str

            if current_time:
                if previous_time:
                    time_diff = (current_time - previous_time).total_seconds()
                    diff_cell = ws.cell(row=row, column=5, value=time_diff)
                    diff_cell.number_format = '0.00'
                previous_time = current_time

            # Step 4: 檢查第3欄是否含有 "offline"
            col3_val = str(ws.cell(row=row, column=3).value).lower()
            if 'offline' in col3_val:
                offline_count += 1
            # 同時計算 "connected" 次數與行數 (若需要)
            if 'connected' in col3_val:
                connected_count += col3_val.count('connected')
                connected_rows += 1

        offline_counts[ws.title] = offline_count
        connected_counts[ws.title] = (connected_count, connected_rows)

    # Step 5: 填寫 Summary 工作表
    if 'summary' not in wb.sheetnames:
        summary_ws = wb.create_sheet('summary', 0)
    else:
        summary_ws = wb['summary']
    
    summary_ws.cell(row=1, column=1, value='File Name')
    summary_ws.cell(row=1, column=2, value='Offline Count')
    summary_ws.cell(row=1, column=3, value='Connected Count')
    summary_ws.cell(row=1, column=4, value='Connected Rows')

    row_index = 2
    for ws_title in offline_counts.keys():
        summary_ws.cell(row=row_index, column=1, value=ws_title)
        summary_ws.cell(row=row_index, column=2, value=offline_counts[ws_title])
        summary_ws.cell(row=row_index, column=3, value=connected_counts[ws_title][0])
        summary_ws.cell(row=row_index, column=4, value=connected_counts[ws_title][1])
        row_index += 1

    # Step 6: 設置所有工作表的字體、對齊和邊框，並自動調整欄寬
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.font = font
                cell.alignment = alignment
                cell.border = border
        for col in ws.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

    # 儲存修改後的檔案 (在原檔名後加 _final)
    new_file_path = file_path.replace('.xlsx', '_final.xlsx')
    wb.save(new_file_path)
    print(f"Processed file saved as {new_file_path}")

# 遍歷資料夾中的所有 Excel 檔案
for file_name in os.listdir(folder_path):
    if file_name.endswith('.xlsx'):
        file_path = os.path.join(folder_path, file_name)
        process_excel(file_path)
