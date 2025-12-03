import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# 設定資料夾路徑
folder_path = r'C:\_python 2024\2024final'

# -------------------- 樣式設定 --------------------
font = Font(name='新細明體', size=12)
alignment = Alignment(horizontal='center', vertical='center')
border = Border(
    left=Side(border_style='thin'),
    right=Side(border_style='thin'),
    top=Side(border_style='thin'),
    bottom=Side(border_style='thin')
)
header_fill = PatternFill(start_color='D0D0D0', end_color='D0D0D0', fill_type='solid')


# -------------------- CSV 轉 Excel --------------------
def convert_csv_to_excel(file_name):
    csv_file_path = os.path.join(folder_path, file_name)
    excel_file_path = os.path.join(folder_path, file_name.replace('.csv', '.xlsx'))

    df = pd.read_csv(csv_file_path)
    df.to_excel(excel_file_path, sheet_name='sheet1', index=False)
    print(f'{file_name} 轉換完成為 {os.path.basename(excel_file_path)}')
    return excel_file_path


# -------------------- 在 sheet1 前面插入欄位 Data select --------------------
def add_column_and_fill_data(file_path):
    workbook = load_workbook(file_path)
    if 'sheet1' in workbook.sheetnames:
        sheet = workbook['sheet1']

        # 在第一欄前插入新的欄位
        sheet.insert_cols(1)

        # 在 A1 儲存格填上 "Data select"
        sheet['A1'] = 'Data select'

        # 從 0 開始計數，並在第二欄有資料的儲存格前填上相對應的數字
        counter = 0
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=2).value is not None:  # 檢查第二欄是否有資料
                sheet.cell(row=row, column=1).value = counter
                counter += 1

        workbook.save(file_path)
        print(f'{os.path.basename(file_path)} 加入 Data select 完成')
    else:
        print(f"{os.path.basename(file_path)} 沒有找到 'sheet1' 工作表")


# -------------------- 產生重複與缺漏時間的工作表 --------------------
def process_duplicates_and_missing(file_path):
    df = pd.read_excel(file_path, sheet_name='sheet1')
    # 假設時間欄位在第 5 欄
    df['Time'] = pd.to_datetime(df.iloc[:, 4], format='%m/%d/%Y %H:%M:%S.%f')

    total_records = len(df)
    duplicates = df[df.duplicated(subset=['Time'], keep=False)]
    duplicate_count = len(duplicates)

    df = df.sort_values(by='Time').reset_index(drop=True)
    time_diffs = df['Time'].diff().dt.total_seconds().fillna(1)

    missing_times_list = []
    for i in range(1, len(df)):
        if time_diffs[i] > 1:
            start_time = df['Time'][i - 1]
            end_time = df['Time'][i]
            missing_times_list.append([start_time, end_time])

    missing_times = pd.DataFrame(missing_times_list, columns=['Start Time', 'End Time'])
    missing_count = len(missing_times)

    print(f"檔案: {os.path.basename(file_path)}")
    print(f"總共有 {total_records} 筆資料")
    print(f"時間重複的筆數: {duplicate_count} 筆")
    print(f"時間有缺漏的筆數: {missing_count} 筆")

    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        duplicates.to_excel(writer, sheet_name='重複筆數', index=False)
        missing_times.to_excel(writer, sheet_name='缺漏筆數', index=False)

    # 將原本第一個工作表改名成「原始 data」
    wb = load_workbook(file_path)
    original_sheet_name = wb.sheetnames[0]
    wb[original_sheet_name].title = '原始 data'
    wb.save(file_path)


# -------------------- 在「缺漏筆數」加入 time difference，並套用樣式 --------------------
def process_time_differences_and_style(file_path):
    wb = load_workbook(file_path)

    # 先處理「缺漏筆數」計算 time difference
    if '缺漏筆數' in wb.sheetnames:
        ws = wb['缺漏筆數']

        # 用 DataFrame 重新計算 time difference
        df = pd.DataFrame(ws.values)
        if len(df) > 0:
            df.columns = df.iloc[0]
            df = df[1:]

            if 'Start Time' in df.columns and 'End Time' in df.columns:
                df['Start Time'] = pd.to_datetime(df['Start Time'])
                df['End Time'] = pd.to_datetime(df['End Time'])
                df['time difference'] = (df['End Time'] - df['Start Time']).dt.total_seconds()
            else:
                # 欄位不符合預期時，建立空的 time difference
                df['time difference'] = pd.Series(dtype=float)

            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name='缺漏筆數', index=False)

            # 計算 time difference 欄位的頻數並新增工作表
            if 'time difference' in df.columns:
                time_diff_counts = df['time difference'].value_counts().reset_index()
                time_diff_counts.columns = ['Time Difference (seconds)', 'Frequency']

                with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    time_diff_counts.to_excel(writer, sheet_name='Time Difference Frequency', index=False)

        # 重新載入工作簿以反映最新內容
        wb = load_workbook(file_path)

    # 對所有工作表套用樣式與自動欄寬/列高
    for sheet in wb.worksheets:
        # 套用字型、對齊、框線、底色
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = font
                cell.alignment = alignment
                cell.border = border
                if cell.value is not None:
                    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        # 第一列當標題列加深底色
        first_row = sheet[1]
        for cell in first_row:
            if cell.value is not None:
                cell.fill = header_fill

        # 自動欄寬
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            adjusted_width = max_length + 2
            sheet.column_dimensions[column].width = adjusted_width

        # 自動列高（粗略估計）
        for row in sheet.iter_rows():
            max_height = 15
            for cell in row:
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                    num_lines = (cell_length // 50) + 1
                    height = num_lines * 15
                    if height > max_height:
                        max_height = height
            sheet.row_dimensions[row[0].row].height = max_height

    wb.save(file_path)


# -------------------- 依 deviceid + 日期重新命名檔案 --------------------
def rename_file(file_path):
    workbook = load_workbook(file_path, data_only=True)
    if '原始 data' in workbook.sheetnames:
        sheet = workbook['原始 data']

        # 查找 deviceid 欄位的列號
        deviceid_col = None
        for col in range(1, sheet.max_column + 1):
            header_value = sheet.cell(row=1, column=col).value
            if header_value == 'deviceid':
                deviceid_col = col
                break

        if deviceid_col:
            # 取得 deviceid 和時間值
            deviceid = sheet.cell(row=2, column=deviceid_col).value
            time_value = sheet.cell(row=2, column=5).value  # 假設時間值在第5欄

            if deviceid and time_value:
                # 轉換時間格式到 YYYYMMDD
                date_str = pd.to_datetime(time_value).strftime('%Y%m%d')

                # 生成新檔名
                new_filename = f"{deviceid}_TW_{date_str}.xlsx"
                new_file_path = os.path.join(folder_path, new_filename)

                # 重命名檔案，取代原檔案
                os.rename(file_path, new_file_path)
                print(f"已將 {os.path.basename(file_path)} 重命名為 {new_filename}")
                return new_file_path

    # 如果無法重命名，返回原檔案路徑
    return file_path


# -------------------- 依每個檔案計算統計，回傳一列字典 --------------------
def compute_statistics(file_path):
    """
    回傳格式：
    {
        'MAC': ...,
        '時間差<1s次數': ...,
        '時間差>=1s & <2s次數': ...,
        '時間差>=2s & <3s次數': ...,
        '時間差>=3s & <4s次數': ...,
        '時間差>=4s & <5s次數': ...,
        '時間差>=5s & <60s次數': ...,
        '時間差>=60s次數': ...,
        '最高時間差秒數': ...,
        '資料重複筆數': ...,
        '資料總筆數': ...   # ★ 新增欄位（原始 csv 的總筆數）
    }
    """
    # 先預設全部都是 0 / None
    stats = {
        'MAC': '',
        '時間差<1s次數': 0,
        '時間差>=1s & <2s次數': 0,
        '時間差>=2s & <3s次數': 0,
        '時間差>=3s & <4s次數': 0,
        '時間差>=4s & <5s次數': 0,
        '時間差>=5s & <60s次數': 0,
        '時間差>=60s次數': 0,
        '最高時間差秒數': None,
        '資料重複筆數': 0,
        '資料總筆數': 0
    }

    wb = load_workbook(file_path, data_only=True)

    # ---- 取得 MAC (deviceid) 與資料總筆數 ----
    if '原始 data' in wb.sheetnames:
        sheet = wb['原始 data']

        # 找 deviceid 欄位
        deviceid_col = None
        for col in range(1, sheet.max_column + 1):
            header_value = sheet.cell(row=1, column=col).value
            if header_value == 'deviceid':
                deviceid_col = col
                break

        if deviceid_col:
            mac_value = sheet.cell(row=2, column=deviceid_col).value
            if mac_value is not None:
                stats['MAC'] = str(mac_value)

        # 資料總筆數 = 除標題列以外的列數（假設沒有多餘空白列）
        data_row_count = max(sheet.max_row - 1, 0)
        stats['資料總筆數'] = int(data_row_count)

    # ---- 從「缺漏筆數」取 time difference 做區間計數 ----
    if '缺漏筆數' in wb.sheetnames:
        try:
            df_missing = pd.read_excel(file_path, sheet_name='缺漏筆數')
            if 'time difference' in df_missing.columns:
                td = pd.to_numeric(df_missing['time difference'], errors='coerce').dropna()

                if not td.empty:
                    stats['時間差<1s次數'] = int((td <= 1).sum())
                    stats['時間差>=1s & <2s次數'] = int(((td >= 1) & (td < 2)).sum())
                    stats['時間差>=2s & <3s次數'] = int(((td >= 2) & (td < 3)).sum())
                    stats['時間差>=3s & <4s次數'] = int(((td >= 3) & (td < 4)).sum())
                    stats['時間差>=4s & <5s次數'] = int(((td >= 4) & (td < 5)).sum())
                    stats['時間差>=5s & <60s次數'] = int(((td >= 5) & (td < 60)).sum())
                    stats['時間差>=60s次數'] = int((td >= 60).sum())
                    stats['最高時間差秒數'] = float(td.max())
        except Exception as e:
            print(f"{os.path.basename(file_path)} 計算缺漏筆數統計時發生錯誤：{e}")

    # ---- 從「重複筆數」計算資料重複筆數 ----
    if '重複筆數' in wb.sheetnames:
        try:
            df_dup = pd.read_excel(file_path, sheet_name='重複筆數')
            stats['資料重複筆數'] = int(len(df_dup))
        except Exception as e:
            print(f"{os.path.basename(file_path)} 讀取重複筆數時發生錯誤：{e}")

    return stats


# -------------------- 主流程 --------------------
if __name__ == '__main__':
    # ① 先把所有 CSV 轉成 Excel
    for file_name in os.listdir(folder_path):
        if file_name.lower().endswith('.csv'):
            convert_csv_to_excel(file_name)

    # 用來累積每一個 MAC 的統計資料
    summary_rows = []

    # ② 處理資料、產生缺漏/重複、時間差、改檔名，同時蒐集統計資料
    for file_name in os.listdir(folder_path):
        if file_name.lower().endswith('.xlsx'):
            # 先排除之後可能產生的統計總表，以免被再次處理
            if file_name == 'MAC_time_diff_summary.xlsx':
                continue

            excel_file_path = os.path.join(folder_path, file_name)

            # 依原本流程處理
            add_column_and_fill_data(excel_file_path)
            process_duplicates_and_missing(excel_file_path)
            process_time_differences_and_style(excel_file_path)
            new_excel_file_path = rename_file(excel_file_path)

            # 針對處理/改名後的檔案計算統計，加入 summary
            stats = compute_statistics(new_excel_file_path)
            summary_rows.append(stats)

    # ③ 產生獨立統計總表 Excel（含第 11 欄：資料總筆數）
    if summary_rows:
        summary_df = pd.DataFrame(summary_rows, columns=[
            'MAC',
            '時間差<1s次數',
            '時間差>=1s & <2s次數',
            '時間差>=2s & <3s次數',
            '時間差>=3s & <4s次數',
            '時間差>=4s & <5s次數',
            '時間差>=5s & <60s次數',
            '時間差>=60s次數',
            '最高時間差秒數',
            '資料重複筆數',
            '資料總筆數'
        ])

        summary_file_path = os.path.join(folder_path, 'MAC_time_diff_summary.xlsx')
        summary_df.to_excel(summary_file_path, index=False)
        print(f"已產生統計總表：{summary_file_path}")
    else:
        print("沒有任何資料可產生統計總表。")
        
#-------------------------------------------------------------------------------
# 把底下的程式執行完後,需再另外產生一獨立excel表格,此表格欄位範例如上傳的截圖,請提供修改後的Python程式碼
# 第1個欄位(MAC): 此MAC的資料,即為執行的csv檔裡的deviceid欄位下的資料
# 第2個欄位(時間差<1s次數): 計算你執行底下程式後,產出的各個檔案裡被命名為缺漏筆數的工作表,統計其第3欄time difference下的時間,如果數字是<=1的,數量共幾個
# 第3個欄位(時間差>=1s & <2s次數): 計算你執行底下程式後,產出的各個檔案裡被命名為缺漏筆數的工作表,統計其第3欄time difference下的時間,如果數字是>=1 & <2的,數量共幾個
# 第4個欄位(時間差>=2s & <3s次數): 計算你執行底下程式後,產出的各個檔案裡被命名為缺漏筆數的工作表,統計其第3欄time difference下的時間,如果數字是>=2& <3的,數量共幾個
# 第5個欄位(時間差>=3s & <4s次數): 計算你執行底下程式後,產出的各個檔案裡被命名為缺漏筆數的工作表,統計其第3欄time difference下的時間,如果數字是>=3& <4的,數量共幾個?
# 第6個欄位(時間差>=4s & <5s次數): 計算你執行底下程式後,產出的各個檔案裡被命名為缺漏筆數的工作表,統計其第3欄time difference下的時間,如果數字是>=4& <5的,數量共幾個?
# 第7個欄位(時間差>=5s<=60s次數): 計算你執行底下程式後,產出的各個檔案裡被命名為缺漏筆數的工作表,統計其第3欄time difference下的時間,如果數字是>=5& <60的,數量共幾個?
# 第8個欄位(時間差>=60s次數): 計算你執行底下程式後,產出的各個檔案裡被命名為缺漏筆數的工作表,統計其第3欄time difference下的時間,如果數字是>=60的,數量共幾個?
# 第9個欄位(最高時間差秒數): 計算你執行底下程式後,產出的各個檔案裡被命名為缺漏筆數的工作表,統計其第3欄time difference下的時間,取出其最大值是
# 第10個欄位(資料重複筆數): 計算你執行底下程式後,產出的各個檔案裡被命名為重複筆數的工作表,統計總共幾筆資料 
# 第11個欄位(資料總筆數): 計算你執行的原始csv檔,每個檔案裡的原始資料是幾筆 

