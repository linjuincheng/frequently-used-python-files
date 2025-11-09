


import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 設定資料夾路徑
folder_path = r'C:\_python 2024\2024final'

# 取得資料夾中的所有 Excel 檔案
excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]

# 逐個處理 Excel 檔案
for file in excel_files:
    file_path = os.path.join(folder_path, file)
    
    # 使用 openpyxl 加載工作簿
    wb = load_workbook(file_path)
    
    # 建立一個 summary 工作表，插入在最前面
    if 'summary' in wb.sheetnames:
        ws_summary = wb['summary']
    else:
        ws_summary = wb.create_sheet(title='summary', index=0)
    
    # 針對所有工作表檢查 offline 字串，忽略 sheet1 和 sheet2
    offline_counts = []
    for sheet_name in wb.sheetnames:
        if sheet_name not in ['sheet1', 'sheet2', 'summary']:
            ws = wb[sheet_name]
            
            # 檢查第三欄的每一列資料是否包含 "offline"
            offline_count = 0
            for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
                cell_value = str(row[0].value).lower() if row[0].value else ''
                if 'offline' in cell_value:
                    offline_count += 1
            
            # 記錄該工作表的名稱和斷線次數
            offline_counts.append((sheet_name, offline_count))
    
    # 將結果寫入 summary 工作表
    ws_summary.append(['Worksheet Name', 'Offline Count'])  # 標題行
    for sheet_name, count in offline_counts:
        ws_summary.append([sheet_name, count])
    
    # 儲存檔案
    wb.save(file_path)

print("資料處理完成！")





# 多個excel檔(其預設路徑放在C:\_python 2024\2024final), 
# 1.	忽略sheet1及sheet2, 檢查其它每一個工作表的第三個欄位的每一列資料, 看是否資料裡帶有offline
# 2.	計算有每個工作表有多少列的資料帶有offline,則為斷線次數
# 3.	在excel檔所有工作前最前面產出名為summary的工作表, 把各工作表的檔名放在summary工作表的第1欄, 計算出的斷線次數放在summary 工作表的第2欄
# 4.	用Python語法


# 說明：
# 忽略 sheet1 和 sheet2：程式會略過名為 sheet1 和 sheet2 的工作表，只檢查其餘的工作表。
# 檢查 "offline" 字串：程式會逐行檢查每個工作表的第三欄，並檢測該欄的資料中是否包含 "offline" 字串（不區分大小寫）。
# 斷線次數計算：計算每個工作表中 "offline" 字串出現的次數，並記錄在 summary 工作表中。
# 在 summary 中寫入結果：將工作表名稱與相應的斷線次數寫入 summary 工作表，該工作表會被插入到最前面。
# 檔案處理流程：
# 程式會自動處理資料夾中的所有 Excel 檔案。
# 結果會儲存在同一個 Excel 檔案的 summary 工作表中，供後續檢查使用。

