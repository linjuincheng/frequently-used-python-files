


import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

# 設定資料夾路徑
folder_path = r'C:\_python 2024\2024final'

# 取得資料夾中的所有 Excel 檔案
excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]

# 定義樣式
font = Font(name='新細明體', size=12)
alignment = Alignment(horizontal='center', vertical='center')
border = Border(left=Side(border_style='thin'), 
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin'))

# 逐個處理 Excel 檔案
for file in excel_files:
    file_path = os.path.join(folder_path, file)
    
    # 讀取 Excel 檔案的 sheet1
    df = pd.read_excel(file_path, sheet_name='sheet1')
    
    # 計算時間差並新增一列，如果存在Time欄位
    if 'Time' in df.columns:
        df['Time'] = pd.to_datetime(df['Time'])  # 確保 'Time' 欄位為時間格式
        df['Time_Diff'] = df['Time'].diff().shift(-1).fillna(pd.Timedelta(seconds=0))
        df['Time_Diff'] = df['Time_Diff'].dt.total_seconds()  # 將時間差轉換為秒數
    
    # 複製資料到 sheet2，不做時間轉換
    df_sheet2 = df.copy()
    
    # 打開現有的Excel檔案
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        # 將 df_sheet2 寫入 sheet2
        df_sheet2.to_excel(writer, sheet_name='sheet2', index=False)
        
        # 根據 MAC address 分組
        for mac_address, group in df_sheet2.groupby(df_sheet2.iloc[:, 0]):
            # 將資料寫入新的工作表，名稱為 MAC address
            group.to_excel(writer, sheet_name=mac_address, index=False)
    
    # 打開Excel檔案以便進行樣式調整
    wb = load_workbook(file_path)
    
    # 遍歷所有工作表並應用樣式
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        
        # 調整每個儲存格的字體、對齊和邊框
        for row in ws.iter_rows():
            for cell in row:
                cell.font = font
                cell.alignment = alignment
                cell.border = border
        
        # 自動調整欄寬
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # 取得欄位字母
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
        
        # 自動調整列高
        for row in ws.iter_rows():
            max_height = 0
            for cell in row:
                if cell.value:
                    cell_height = len(str(cell.value)) // 20 + 1
                    max_height = max(max_height, cell_height)
            ws.row_dimensions[row[0].row].height = max_height * 15
    
    # 建立一個 summary 工作表，插入在最前面
    if 'summary' in wb.sheetnames:
        ws_summary = wb['summary']
    else:
        ws_summary = wb.create_sheet(title='summary', index=0)
    
    wb.save(file_path)
