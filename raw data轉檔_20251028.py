import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# 設定資料夾路徑
folder_path = r'C:\_python 2024\2024final'

# ===== 樣式設定（保留；若不需要樣式，可整段刪除並略過 apply_style 的呼叫）=====
font = Font(name='新細明體', size=12)
alignment = Alignment(horizontal='center', vertical='center')
border = Border(left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin'))
header_fill = PatternFill(start_color='D0D0D0', end_color='D0D0D0', fill_type='solid')
# ===========================================================================

# CSV 檔案轉換為只有「原始 data」工作表的 Excel
def convert_csv_to_excel(file_name):
    csv_file_path = os.path.join(folder_path, file_name)
    excel_file_path = os.path.join(folder_path, file_name.replace('.csv', '.xlsx'))

    # 讀入 CSV（有需要可自訂 encoding / sep）
    df = pd.read_csv(csv_file_path)
    # 直接寫成「原始 data」工作表
    df.to_excel(excel_file_path, sheet_name='原始 data', index=False)
    print(f'{file_name} 轉換完成為 {os.path.basename(excel_file_path)}')
    return excel_file_path

# 在「原始 data」插入第 1 欄並填入遞增編號
def add_column_and_fill_data(file_path):
    workbook = load_workbook(file_path)
    if '原始 data' in workbook.sheetnames:
        sheet = workbook['原始 data']

        # 在第一欄前插入新的欄位
        sheet.insert_cols(1)

        # A1 放標題
        sheet['A1'] = 'Data select'

        # 從 0 開始計數，若第二欄（原來的第 1 欄）有資料就填號
        counter = 0
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=2).value is not None:
                sheet.cell(row=row, column=1).value = counter
                counter += 1

        workbook.save(file_path)
        print(f'{os.path.basename(file_path)}（插入欄位與編號）處理完成')
    else:
        print(f"{os.path.basename(file_path)} 沒有找到 '原始 data' 工作表")

#（可選）只對現存工作表（此版本只有「原始 data」）套樣式與自動欄寬列高
def apply_style(file_path):
    wb = load_workbook(file_path)
    for sheet in wb.worksheets:
        # 套用字型、置中、框線與底色（一般儲存格白底）
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = font
                cell.alignment = alignment
                cell.border = border
                if cell.value is not None:
                    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        # 表頭列底色
        first_row = sheet[1]
        for cell in first_row:
            if cell.value is not None:
                cell.fill = header_fill

        # 自動欄寬（粗略做法：依內容長度）
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            adjusted_width = max_length + 2
            sheet.column_dimensions[column].width = adjusted_width

        # 自動列高（粗略：每 50 字一倍高度）
        for row in sheet.iter_rows():
            max_height = 15
            for cell in row:
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                    num_lines = (cell_length // 50) + 1
                    height = num_lines * 15
                    if height > max_height:
                        max_height = height
            sheet.row_dimensions[row[0].row].height = max_height

    wb.save(file_path)
    print(f'{os.path.basename(file_path)} 樣式套用完成')

# 依「原始 data」中的 deviceid 與第 5 欄時間重命名
def rename_file(file_path):
    workbook = load_workbook(file_path, data_only=True)
    if '原始 data' in workbook.sheetnames:
        sheet = workbook['原始 data']

        # 找出 deviceid 欄位位置
        deviceid_col = None
        for col in range(1, sheet.max_column + 1):
            header_value = sheet.cell(row=1, column=col).value
            if header_value == 'deviceid':
                deviceid_col = col
                break

        if deviceid_col:
            deviceid = sheet.cell(row=2, column=deviceid_col).value
            time_value = sheet.cell(row=2, column=5).value  # 假設時間在第 5 欄

            if deviceid and time_value:
                # 轉成 YYYYMMDD
                date_str = pd.to_datetime(time_value).strftime('%Y%m%d')
                new_filename = f"{deviceid}_TW_{date_str}.xlsx"
                new_file_path = os.path.join(folder_path, new_filename)

                # 避免同名覆蓋：若已存在就先刪除或改名（這裡選擇覆蓋）
                if os.path.exists(new_file_path):
                    os.remove(new_file_path)

                os.rename(file_path, new_file_path)
                print(f"已將 {os.path.basename(file_path)} 重命名為 {new_filename}")
                return new_file_path

    return file_path  # 無法重命名則回傳原路徑

# =================== 主流程 ===================

# 1) 轉換所有 CSV 成只有「原始 data」工作表的 Excel
for file_name in os.listdir(folder_path):
    if file_name.lower().endswith('.csv'):
        convert_csv_to_excel(file_name)

# 2) 對所有 Excel 進行「插欄位 + 編號」→（可選）套樣式 → 重命名
for file_name in os.listdir(folder_path):
    if file_name.lower().endswith('.xlsx'):
        excel_file_path = os.path.join(folder_path, file_name)
        add_column_and_fill_data(excel_file_path)
        # 若不想套樣式，將下一行註解掉
        apply_style(excel_file_path)
        rename_file(excel_file_path)

