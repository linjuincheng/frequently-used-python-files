# 要把數個excel裡所有工作表中的字體變更為新細明體，字體大小變更為12，所有儲存格的內容都變更為置中對齊，並且要依儲存格內容資料長度來自動調整列高和自動調整欄寬(要讓儲存格完全顯示),且第一列有內容的儲存格, 佈景主題顏色都改這顏色(白色,背景1,較深15%),且儲存格內有內容的儲存格一律加上框線 , 輸出的檔名用原檔名, 用Python語法


import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# 定義樣式
font = Font(name='新細明體', size=12)
alignment = Alignment(horizontal='center', vertical='center')
border = Border(left=Side(border_style='thin'), 
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin'))

# 標題列顏色設定
header_fill = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')  # 白色, 背景1, 較深15%

# 定義處理 Excel 文件的函數
def process_excel(file_path):
    wb = load_workbook(file_path)
    for sheet in wb.worksheets:
        # 設置所有儲存格的字體、對齊方式和邊框
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = font
                cell.alignment = alignment
                cell.border = border
                if cell.value:
                    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        # 設置標題列的背景顏色
        first_row = sheet[1]
        for cell in first_row:
            if cell.value is not None:
                cell.fill = header_fill

        # 自動調整列寬以顯示內容
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            adjusted_width = max_length + 2
            sheet.column_dimensions[column].width = adjusted_width

        # 自動調整行高以顯示內容
        for row in sheet.iter_rows():
            max_height = 15  # 基本行高
            for cell in row:
                if cell.value:
                    # 根據儲存格內容的長度計算行高
                    cell_length = len(str(cell.value))
                    num_lines = (cell_length // 50) + 1  # 每行最多顯示 50 個字符
                    height = num_lines * 15  # 假設每行的高度為 15 點
                    if height > max_height:
                        max_height = height
            sheet.row_dimensions[row[0].row].height = max_height

    # 保存更改，使用原檔名
    wb.save(file_path)

# 設定需要處理的目錄
directory = r'C:\_python 2024\2024final'

# 遍歷目錄中的所有 Excel 文件並處理
for filename in os.listdir(directory):
    if filename.endswith('.xlsx'):
        file_path = os.path.join(directory, filename)
        process_excel(file_path)
