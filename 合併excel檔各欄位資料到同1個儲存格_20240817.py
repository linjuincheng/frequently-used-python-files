

import openpyxl
import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

def process_excel_file(file_path):
    # 打開 Excel 文件
    wb = openpyxl.load_workbook(file_path)
    sheet1 = wb['sheet1']
    
    # 新建 sheet2
    if 'sheet2' not in wb.sheetnames:
        wb.create_sheet('sheet2')
    sheet2 = wb['sheet2']
    
    # 設定字體為'新細明體', 大小為12
    font = Font(name='新細明體', size=12)
    
    row_counter = 1  # 用於 sheet2 的行計數
    
    # 遍歷 sheet1，每2列起的每一列處理
    for row_index, row in enumerate(sheet1.iter_rows(min_row=2), start=2):
        if row_index % 2 == 0:  # 每隔一行處理
            combined_data = []
            
            # 將第1到第15個欄位的資料用空格合併
            for cell in row[:25]:  # 只取前25個欄位
                combined_data.append(str(cell.value) if cell.value is not None else '')
            
            # 合併資料並留一個空格隔開
            combined_str = ' '.join(combined_data)
            
            # 將合併後的資料寫入 sheet2 的第1欄
            cell = sheet2.cell(row=row_counter, column=1, value=combined_str)
            cell.font = font  # 設定字體
            
            row_counter += 1
    
    # 自動調整欄寬
    for col in sheet2.columns:
        max_length = 0
        column = col[0].column  # 列的編號 (數字)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet2.column_dimensions[get_column_letter(column)].width = adjusted_width

    # 自動調整列高
    for row in sheet2.iter_rows():
        sheet2.row_dimensions[row[0].row].height = 20  # 調整列高為20，可以根據需求調整
    
    # 保存文件
    wb.save(file_path)
    print(f"處理完成並保存至 {file_path}")

def process_multiple_excel_files(folder_path):
    # 處理該文件夾下的所有 Excel 文件
    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx") or filename.endswith(".xlsm"):
            file_path = os.path.join(folder_path, filename)
            process_excel_file(file_path)

# 指定存放多個 Excel 檔案的文件夾路徑
folder_path = 'C:\\_python 2024\\2024final' 
process_multiple_excel_files(folder_path)



# 多個excel檔, 每個excel檔的sheet1的每2列起的每一列如果其第1到第15個欄位裡有資料或者空白, 要把每一列的15個欄位裡的資料留一個空格後合併在1個欄位裡, 字形統一為name='新細明體', size=12 ,且自動調整列高及自動調整欄高,然後將合併的資料產出一個sheet2, 輸出的excel檔名採用原檔名, 用Python語法

# 3. 說明：
# process_excel_file 函數：處理單個 Excel 文件。對每 2 列起的每一列，將前 15 個欄位的資料合併，並用空格隔開。合併後的資料寫入到新的 sheet2 中。字體統一為 '新細明體'、字型大小為 12，並自動調整欄寬與列高。
# process_multiple_excel_files 函數：遍歷指定文件夾中的所有 .xlsx 或 .xlsm 文件，並對每個文件進行處理。
# 4. 使用方式：
# 將要處理的 Excel 文件放入同一個文件夾中。
# 將程式中的 folder_path 修改為包含這些 Excel 文件的文件夾路徑。
# 執行程式，程式會自動處理每個 Excel 文件，並將結果輸出到同一個文件中的 sheet2。
# 此程式將根據您的需求來自動化處理 Excel 文件並保存修改。