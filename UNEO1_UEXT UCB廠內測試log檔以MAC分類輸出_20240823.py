

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import pytz

# 設定資料夾路徑和台灣時區
folder_path = r'C:\_python 2024\2024final'
taiwan_tz = pytz.timezone('Asia/Taipei')

# 取得資料夾中的所有 Excel 檔案
excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]

# 定義樣式
font = Font(name='新細明體', size=12)
alignment = Alignment(horizontal='center', vertical='center')
border = Border(left=Side(border_style='thin'), 
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin'))

def convert_utc_to_taiwan_time(utc_time):
    """將 UTC 時間轉換為台灣時間，並移除時區資訊"""
    if isinstance(utc_time, pd.Timestamp):
        utc_time = utc_time.strftime('%Y-%m-%d %H:%M:%S.%f')
    try:
        utc_time = datetime.strptime(utc_time, '%Y-%m-%d %H:%M:%S.%f')
    except ValueError:
        utc_time = datetime.strptime(utc_time, '%Y-%m-%d %H:%M:%S')
    utc_time = pytz.utc.localize(utc_time)
    taiwan_time = utc_time.astimezone(taiwan_tz).replace(tzinfo=None)
    return taiwan_time

# 逐個處理 Excel 檔案
for file in excel_files:
    file_path = os.path.join(folder_path, file)
    
    # 讀取 Excel 檔案的 sheet1
    df = pd.read_excel(file_path, sheet_name='sheet1')
    
    # 複製資料到 sheet2 並轉換第4欄UTC時間為台灣時間
    df_sheet2 = df.copy()
    df_sheet2.iloc[:, 3] = df_sheet2.iloc[:, 3].apply(convert_utc_to_taiwan_time)
    
    # 打開現有的Excel檔案
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        # 將 df_sheet2 寫入 sheet2
        df_sheet2.to_excel(writer, sheet_name='sheet2', index=False)
        
        # 根據 MAC address 分組
        for mac_address, group in df_sheet2.groupby(df_sheet2.iloc[:, 0]):
            # 將資料寫入新的工作表，名稱為 MAC address
            group.to_excel(writer, sheet_name=mac_address, index=False)
    
    # 打開Excel檔案以便進行樣式調整
    wb = load_workbook(file_path)
    
    # 遍歷所有工作表並應用樣式
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        
        # 調整每個儲存格的字體、對齊和邊框
        for row in ws.iter_rows():
            for cell in row:
                cell.font = font
                cell.alignment = alignment
                cell.border = border
        
        # 自動調整欄寬
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # 取得欄位字母
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
        
        # 自動調整列高
        for row in ws.iter_rows():
            max_height = 0
            for cell in row:
                if cell.value:
                    cell_height = len(str(cell.value)) // 20 + 1
                    max_height = max(max_height, cell_height)
            ws.row_dimensions[row[0].row].height = max_height * 15
    
    # 儲存調整後的檔案
    wb.save(file_path)

print("資料處理完成！")


# 多個excel檔(其預設路徑放在C:\_python 2024\2024final), 其每個excel檔的sheet1裡的儲存格有資料, 總共4個欄位, 第1個欄位是MAC address, 第4欄位為UTC時間 
# 1.	產出新的工作表取名sheet2, 將shee1的資料複製到sheet2並直接把sheet2第4欄位的UTC時間改為台灣時間
# 2.	將sheet2裡1個欄位裡同樣MAC address的篩選出來, 同樣MAC address產出一個新的工作表 , 工作表接續在sheet2之後, 並把工作表的名稱以MAC address的名稱命名(e.g. 80C9555CC3CC)
# 3.	將所有工作表的調整字體及欄位, 如下列程式碼
# font = Font(name='新細明體', size=12)
# alignment = Alignment(horizontal='center', vertical='center')
# border = Border(left=Side(border_style='thin'), 
#                 right=Side(border_style='thin'),
#                 top=Side(border_style='thin'),
#                 bottom=Side(border_style='thin'))
# 用Python語法

# 程式說明：
# 資料處理與轉換：sheet1的資料被複製到sheet2，並將第4欄的UTC時間轉換為台灣時間。
# 分組與產生新工作表：根據相同的MAC address篩選出資料，並產生一個新工作表，名稱為該MAC address。
# 樣式調整：對所有工作表應用字體、對齊和邊框格式。
# 執行此程式後，會在每個Excel檔案中新增sheet2與按MAC address命名的工作表，並將所有工作表的字體和格式設置為指定樣式。

# 執行上述程式碼產出各工作表後, 還是把所有工作表的字體設為新細明體', size=12, 儲存格欄位置中對齊,自動調整列高及自動調整欄高後再輸出