

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
import pytz

# 設定資料夾路徑和台灣時區
folder_path = r'C:\_python 2024\2024final'
taiwan_tz = pytz.timezone('Asia/Taipei')

# 取得資料夾中的所有 Excel 檔案
excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]

# 定義樣式
font = Font(name='新細明體', size=12)
alignment = Alignment(horizontal='center', vertical='center')
border = Border(left=Side(border_style='thin'), 
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin'))

def convert_utc_to_taiwan_time(utc_time):
    """將 UTC 時間轉換為台灣時間，並移除時區資訊"""
    if isinstance(utc_time, pd.Timestamp):
        utc_time = utc_time.strftime('%Y-%m-%d %H:%M:%S.%f')
    try:
        utc_time = datetime.strptime(utc_time, '%Y-%m-%d %H:%M:%S.%f')
    except ValueError:
        utc_time = datetime.strptime(utc_time, '%Y-%m-%d %H:%M:%S')
    utc_time = pytz.utc.localize(utc_time)
    taiwan_time = utc_time.astimezone(taiwan_tz).replace(tzinfo=None)
    return taiwan_time

# 逐個處理 Excel 檔案
for file in excel_files:
    file_path = os.path.join(folder_path, file)
    
    # 讀取 Excel 檔案的 sheet1
    df = pd.read_excel(file_path, sheet_name='sheet1')
    
    # 複製資料到 sheet2 並轉換第4欄UTC時間為台灣時間
    df_sheet2 = df.copy()
    df_sheet2.iloc[:, 3] = df_sheet2.iloc[:, 3].apply(convert_utc_to_taiwan_time)
    
    # 打開現有的Excel檔案
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        # 將 df_sheet2 寫入 sheet2
        df_sheet2.to_excel(writer, sheet_name='sheet2', index=False)
        
        # 根據 MAC address 分組
        for mac_address, group in df_sheet2.groupby(df_sheet2.iloc[:, 0]):
            # 將資料寫入新的工作表，名稱為 MAC address
            group.to_excel(writer, sheet_name=mac_address, index=False)
    
    # 打開Excel檔案以便進行樣式調整
    wb = load_workbook(file_path)
    
    # 遍歷所有工作表並應用樣式
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        
        # 調整每個儲存格的字體、對齊和邊框
        for row in ws.iter_rows():
            for cell in row:
                cell.font = font
                cell.alignment = alignment
                cell.border = border
        
        # 自動調整欄寬
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # 取得欄位字母
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
        
        # 自動調整列高
        for row in ws.iter_rows():
            max_height = 0
            for cell in row:
                if cell.value:
                    cell_height = len(str(cell.value)) // 20 + 1
                    max_height = max(max_height, cell_height)
            ws.row_dimensions[row[0].row].height = max_height * 15
    
    # 建立一個 summary 工作表，插入在最前面
    if 'summary' in wb.sheetnames:
        ws_summary = wb['summary']
    else:
        ws_summary = wb.create_sheet(title='summary', index=0)
    
    # 針對所有工作表檢查 offline 字串，忽略 sheet1 和 sheet2
    offline_counts = []
    for sheet_name in wb.sheetnames:
        if sheet_name not in ['sheet1', 'sheet2', 'summary']:
            ws = wb[sheet_name]
            
            # 檢查第三欄的每一列資料是否包含 "offline"
            offline_count = 0
            for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
                cell_value = str(row[0].value).lower() if row[0].value else ''
                if 'offline' in cell_value:
                    offline_count += 1
            
            # 記錄該工作表的名稱和斷線次數
            offline_counts.append((sheet_name, offline_count))
    
    # 將結果寫入 summary 工作表
    ws_summary.append(['Worksheet Name', 'Offline Count'])  # 標題行
    for sheet_name, count in offline_counts:
        ws_summary.append([sheet_name, count])
    
    # 儲存調整後的檔案
    wb.save(file_path)

print("資料處理完成！")



# 程式碼說明：
# 資料處理：程式會逐一讀取 Excel 檔案，並將 sheet1 的資料複製到 sheet2，轉換其中的 UTC 時間為台灣時間，接著根據 MAC 地址分組，將每個組別的資料寫入一個新的工作表。

# 格式調整：對每個工作表應用字體、對齊、邊框等樣式，並自動調整欄寬和列高。

# 新增 summary 工作表：在所有工作表處理完成後，程式會檢查每個工作表的第三欄是否包含 "offline" 字串，並將每個工作表中 "offline" 的次數記錄到 summary 工作表中。

# 最終結果：
# 資料被整理並分組。
# 所有工作表的資料格式被調整。
# summary 工作表中記錄了各個工作表中 "offline" 的次數。

