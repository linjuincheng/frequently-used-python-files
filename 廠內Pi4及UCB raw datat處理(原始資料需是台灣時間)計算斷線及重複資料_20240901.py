import os
import pandas as pd

# 設定資料夾路徑
folder_path = 'C:\\_python 2024\\2024final'

# CSV檔案轉換為Excel檔案
def convert_csv_to_excel(file_name):
    csv_file_path = os.path.join(folder_path, file_name)
    excel_file_path = os.path.join(folder_path, file_name.replace('.csv', '.xlsx'))
    
    df = pd.read_csv(csv_file_path)
    df.to_excel(excel_file_path, sheet_name='sheet1', index=False)
    print(f'{file_name} 轉換完成為 {os.path.basename(excel_file_path)}')
    return excel_file_path

# 轉換所有CSV檔案
for file_name in os.listdir(folder_path):
    if file_name.endswith('.csv'):
        convert_csv_to_excel(file_name)




import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# 設定資料夾路徑
folder_path = 'C:\\_python 2024\\2024final'

# 定義樣式
font = Font(name='新細明體', size=12)
alignment = Alignment(horizontal='center', vertical='center')
border = Border(left=Side(border_style='thin'), 
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin'))
header_fill = PatternFill(start_color='D0D0D0', end_color='D0D0D0', fill_type='solid')

# CSV檔案轉換為Excel檔案
def convert_csv_to_excel(file_name):
    csv_file_path = os.path.join(folder_path, file_name)
    excel_file_path = os.path.join(folder_path, file_name.replace('.csv', '.xlsx'))
    
    df = pd.read_csv(csv_file_path)
    df.to_excel(excel_file_path, sheet_name='sheet1', index=False)
    print(f'{file_name} 轉換完成為 {os.path.basename(excel_file_path)}')
    return excel_file_path

# 新增欄位到sheet1並填入資料
def add_column_and_fill_data(file_path):
    workbook = load_workbook(file_path)
    if 'sheet1' in workbook.sheetnames:
        sheet = workbook['sheet1']

        # 在第一欄前插入新的欄位
        sheet.insert_cols(1)

        # 在A1儲存格填上"Data select"
        sheet['A1'] = 'Data select'

        # 從0開始計數，並在第二欄有資料的儲存格前填上相對應的數字
        counter = 0
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=2).value is not None:  # 檢查第二欄是否有資料
                sheet.cell(row=row, column=1).value = counter  # 在第一欄填上數字
                counter += 1

        workbook.save(file_path)
        print(f'{os.path.basename(file_path)} 處理完成')
    else:
        print(f"{os.path.basename(file_path)} 沒有找到 'sheet1' 工作表")

# 處理重複和缺漏的筆數
def process_duplicates_and_missing(file_path):
    df = pd.read_excel(file_path, sheet_name='sheet1')
    df['Time'] = pd.to_datetime(df.iloc[:, 4], format='%m/%d/%Y %H:%M:%S')
    total_records = len(df)
    duplicates = df[df.duplicated(subset=['Time'], keep=False)]
    duplicate_count = len(duplicates)
    df = df.sort_values(by='Time').reset_index(drop=True)
    time_diffs = df['Time'].diff().dt.total_seconds().fillna(1)
    missing_times_list = []
    for i in range(1, len(df)):
        if time_diffs[i] > 1:
            start_time = df['Time'][i-1]
            end_time = df['Time'][i]
            missing_times_list.append([start_time, end_time])
    missing_times = pd.DataFrame(missing_times_list, columns=['Start Time', 'End Time'])
    missing_count = len(missing_times)
    print(f"檔案: {os.path.basename(file_path)}")
    print(f"總共有 {total_records} 筆資料")
    print(f"時間重複的筆數: {duplicate_count} 筆")
    print(f"時間有缺漏的筆數: {missing_count} 筆")

    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        duplicates.to_excel(writer, sheet_name='重複筆數', index=False)
        missing_times.to_excel(writer, sheet_name='缺漏筆數', index=False)

    wb = load_workbook(file_path)
    original_sheet_name = wb.sheetnames[0]
    wb[original_sheet_name].title = '原始 data'
    wb.save(file_path)

# 設置樣式和格式並處理時間差異
def process_time_differences_and_style(file_path):
    wb = load_workbook(file_path)
    if '缺漏筆數' in wb.sheetnames:
        ws = wb['缺漏筆數']
        df = pd.DataFrame(ws.values)
        df.columns = df.iloc[0]
        df = df[1:]
        df['Start Time'] = pd.to_datetime(df['Start Time'])
        df['End Time'] = pd.to_datetime(df['End Time'])
        df['time difference'] = (df['End Time'] - df['Start Time']).dt.total_seconds()
        
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name='缺漏筆數', index=False)

        # 計算 time difference 欄位的頻數並新增工作表
        time_diff_counts = df['time difference'].value_counts().reset_index()
        time_diff_counts.columns = ['Time Difference (seconds)', 'Frequency']

        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            time_diff_counts.to_excel(writer, sheet_name='Time Difference Frequency', index=False)

    wb = load_workbook(file_path)
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = font
                cell.alignment = alignment
                cell.border = border
                if cell.value:
                    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        first_row = sheet[1]
        for cell in first_row:
            if cell.value is not None:
                cell.fill = header_fill
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            adjusted_width = max_length + 2
            sheet.column_dimensions[column].width = adjusted_width
        for row in sheet.iter_rows():
            max_height = 15
            for cell in row:
                if cell.value:
                    cell_length = len(str(cell.value))
                    num_lines = (cell_length // 50) + 1
                    height = num_lines * 15
                    if height > max_height:
                        max_height = height
            sheet.row_dimensions[row[0].row].height = max_height
    wb.save(file_path)

# 重新命名檔案
def rename_file(file_path):
    workbook = load_workbook(file_path, data_only=True)
    if '原始 data' in workbook.sheetnames:
        sheet = workbook['原始 data']

        # 查找 deviceid 欄位的列號
        deviceid_col = None
        for col in range(1, sheet.max_column + 1):
            header_value = sheet.cell(row=1, column=col).value
            if header_value == 'deviceid':
                deviceid_col = col
                break

        if deviceid_col:
            # 取得 deviceid 和時間值
            deviceid = sheet.cell(row=2, column=deviceid_col).value
            time_value = sheet.cell(row=2, column=5).value  # 假設時間值在第5欄

            if deviceid and time_value:
                # 轉換時間格式到 YYYYMMDD
                date_str = pd.to_datetime(time_value).strftime('%Y%m%d')

                # 生成新檔名
                new_filename = f"{deviceid}_TW_{date_str}.xlsx"
                new_file_path = os.path.join(folder_path, new_filename)

                # 重命名檔案，取代原檔案
                os.rename(file_path, new_file_path)
                print(f"已將 {os.path.basename(file_path)} 重命名為 {new_filename}")
                return new_file_path

    return file_path  # 如果無法重命名，返回原檔案路徑

# 對所有Excel檔案執行處理
for file_name in os.listdir(folder_path):
    if file_name.endswith('.xlsx'):
        excel_file_path = os.path.join(folder_path, file_name)
        add_column_and_fill_data(excel_file_path)
        process_duplicates_and_missing(excel_file_path)
        process_time_differences_and_style(excel_file_path)
        rename_file(excel_file_path)
        

