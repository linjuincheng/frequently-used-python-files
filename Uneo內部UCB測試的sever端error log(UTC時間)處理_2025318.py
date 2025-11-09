import os
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# 設定資料夾路徑 (請確認此路徑存在)
folder_path = r'C:\_python 2024\2024final'

# 定義共用樣式
font = Font(name='新細明體', size=12)
alignment = Alignment(horizontal='center', vertical='center')
border = Border(left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin'))

def utc_to_taiwan(utc_str):
    """將 UTC 時間轉換為台灣時間"""
    try:
        utc_time = datetime.strptime(utc_str, '%Y-%m-%d %H:%M:%S.%f')
    except ValueError:
        try:
            utc_time = datetime.strptime(utc_str, '%Y-%m-%d %H:%M:%S')
        except ValueError:
            raise ValueError(f"Unsupported time format: {utc_str}")
    taiwan_time = utc_time + timedelta(hours=8)
    return taiwan_time.strftime('%Y-%m-%d %H:%M:%S')

def auto_adjust(ws):
    """自動調整欄寬與列高 (模擬 Excel 裡 CTL+A→格式→自動調整)"""
    # 調整欄寬
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                try:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
                except Exception as e:
                    pass
        ws.column_dimensions[col_letter].width = max_length + 2
    # 調整列高
    for row in ws.iter_rows():
        max_height = 0
        for cell in row:
            if cell.value:
                height = len(str(cell.value).split('\n'))
                if height > max_height:
                    max_height = height
        ws.row_dimensions[row[0].row].height = (max_height + 2) * 15

def apply_formatting(ws):
    """設定儲存格字型、置中、加邊框，再自動調整欄寬及列高"""
    for row in ws.iter_rows():
        for cell in row:
            cell.font = font
            cell.alignment = alignment
            cell.border = border
    auto_adjust(ws)

def process_excel(file_path):
    """
    處理原始 Excel 檔案：
      1. 讀取 Sheet1 (或 sheet1)
      2. 建立 sheet2，複製資料並將第4欄(UTC)轉換為台灣時間
      3. 根據 sheet2 第一欄(MAC)建立各獨立工作表
      4. 計算時間差、統計連線/斷線資料
      5. 建立 summary 工作表
      6. 將 sheet1 改名為 Log(UTC)，sheet2 改名為 Log(TW)
      7. 為所有工作表設定字型、置中、邊框與自動調整欄寬列高
      8. 儲存檔案為 _final.xlsx
    """
    wb = load_workbook(file_path)

    # Step 0: 讀取原始資料表 (Sheet1 或 sheet1)
    if 'Sheet1' in wb.sheetnames:
        sheet1 = wb['Sheet1']
    elif 'sheet1' in wb.sheetnames:
        sheet1 = wb['sheet1']
    else:
        print(f"Neither 'Sheet1' nor 'sheet1' found in {file_path}")
        return

    # Step 1: 建立新的工作表 sheet2，複製 sheet1 資料，並將第4欄(UTC時間)轉為台灣時間
    if 'sheet2' in wb.sheetnames:
        sheet2 = wb['sheet2']
    else:
        sheet2 = wb.create_sheet('sheet2')
    # 複製 sheet1 所有列 (包含標題列)
    for row in sheet1.iter_rows(values_only=True):
        sheet2.append(row)
    # 將 sheet2 的第4欄(UTC)轉換為台灣時間 (從第二列開始)
    for row in sheet2.iter_rows(min_row=2, min_col=4, max_col=4):
        cell = row[0]
        if isinstance(cell.value, str):
            try:
                cell.value = utc_to_taiwan(cell.value)
            except ValueError:
                pass

    # Step 2: 根據 sheet2 第一欄(MAC address)建立工作表，
    #         將相同 MAC 的資料複製至同一個新工作表 (工作表名稱即為 MAC address)
    mac_dict = {}
    for row in sheet2.iter_rows(min_row=2, max_col=4, values_only=True):
        mac = row[0]
        if mac not in mac_dict:
            mac_dict[mac] = []
        mac_dict[mac].append(row)
    # 建立 MAC 工作表 (加入至 sheet2 之後)
    for mac, rows in mac_dict.items():
        ws_mac = wb.create_sheet(title=str(mac))
        # 複製標題列 (來自 sheet2)
        for col in range(1, 5):
            cell = ws_mac.cell(row=1, column=col, value=sheet2.cell(row=1, column=col).value)
            cell.font = font
            cell.alignment = alignment
            cell.border = border
        # 寫入資料列
        for data in rows:
            ws_mac.append(data)

    # Step 3 ~ 10: 對除 sheet1、sheet2 及 summary 之外的工作表處理
    exclude_ws = {sheet1.title, sheet2.title, 'summary'}
    stats = {}  # 統計資料：key = MAC 工作表名稱, value = (最高斷線秒數, 斷線>=60s 次數, 斷線<60s 次數, 總斷線次數, 總連線次數)

    for ws in wb.worksheets:
        if ws.title in exclude_ws:
            continue

        # Step 3: 計算時間差 (以第4欄時間計算與前一筆的時間差，單位：秒)，結果放在第5欄 ("Time difference")
        ws.cell(row=1, column=5, value='Time difference')
        previous_time = None
        for row in range(2, ws.max_row + 1):
            time_val = ws.cell(row=row, column=4).value
            current_time = None
            if isinstance(time_val, str):
                try:
                    current_time = datetime.strptime(time_val, '%Y-%m-%d %H:%M:%S')
                except ValueError:
                    try:
                        current_time = datetime.strptime(time_val, '%Y-%m-%d %H:%M:%S.%f')
                    except ValueError:
                        current_time = None
            elif isinstance(time_val, datetime):
                current_time = time_val

            if current_time:
                if previous_time:
                    diff_sec = (current_time - previous_time).total_seconds()
                    diff_cell = ws.cell(row=row, column=5, value=diff_sec)
                    diff_cell.number_format = '0.00'
                previous_time = current_time

        # Step 4: 檢查第3欄是否含有 "offline"，計算總斷線次數
        offline_count = 0
        for row in range(2, ws.max_row + 1):
            if 'offline' in str(ws.cell(row=row, column=3).value).lower():
                offline_count += 1

        # 新增: 計算總連線次數 (第3欄含 "connected")
        connected_count = 0
        for row in range(2, ws.max_row + 1):
            if 'connected' in str(ws.cell(row=row, column=3).value).lower():
                connected_count += 1

        # Step 6: 新增第6欄，標題為 "斷線秒數"
        ws.cell(row=1, column=6, value='斷線秒數')
        # Step 7: 若上一列第3欄含 "offline"，則將本列第5欄 (時間差) 的數值複製到本列第6欄
        for row in range(2, ws.max_row + 1):
            prev_col3 = str(ws.cell(row=row-1, column=3).value).lower() if row-1 >= 2 else ''
            if 'offline' in prev_col3:
                ws.cell(row=row, column=6, value=ws.cell(row=row, column=5).value)
        # Step 8 ~ 10: 統計第6欄數值
        max_disconnect = 0
        disconnect_ge60 = 0
        disconnect_lt60 = 0
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=6).value
            if isinstance(val, (int, float)):
                if val > max_disconnect:
                    max_disconnect = val
                if val >= 60:
                    disconnect_ge60 += 1
                elif val < 60:
                    disconnect_lt60 += 1
        stats[ws.title] = (max_disconnect, disconnect_ge60, disconnect_lt60, offline_count, connected_count)

    # Step 11: 建立 summary 工作表 (放在最前面)
    if 'summary' not in wb.sheetnames:
        summary_ws = wb.create_sheet('summary', 0)
    else:
        summary_ws = wb['summary']
    summary_ws.cell(row=1, column=1, value='MAC')
    summary_ws.cell(row=1, column=2, value='最高斷線秒數')
    summary_ws.cell(row=1, column=3, value='斷線>=60s次數')
    summary_ws.cell(row=1, column=4, value='斷線<60s次數')
    summary_ws.cell(row=1, column=5, value='總斷線次數')
    summary_ws.cell(row=1, column=6, value='總連線次數')
    row_idx = 2
    for ws_name, (max_disc, ge60, lt60, total_offline, total_connected) in stats.items():
        summary_ws.cell(row=row_idx, column=1, value=ws_name)
        summary_ws.cell(row=row_idx, column=2, value=max_disc)
        summary_ws.cell(row=row_idx, column=3, value=ge60)
        summary_ws.cell(row=row_idx, column=4, value=lt60)
        summary_ws.cell(row=row_idx, column=5, value=total_offline)
        summary_ws.cell(row=row_idx, column=6, value=total_connected)
        row_idx += 1

    # Step 12 & 13: 將 sheet1 改名為 Log(UTC)；將 sheet2 改名為 Log(TW)
    sheet1.title = 'Log(UTC)'
    sheet2.title = 'Log(TW)'

    # Step 14: 為所有工作表設定字型、置中、加邊框，並自動調整欄寬及列高
    for ws in wb.worksheets:
        apply_formatting(ws)

    # Step 15: 以原檔名加上 _final 儲存檔案
    new_file_path = file_path.replace('.xlsx', '_final.xlsx')
    wb.save(new_file_path)
    print(f"Processed file saved as {new_file_path}")

def highlight_offline_rows(ws):
    """
    檢查該工作表每一列，
    若該列第3欄儲存格包含 'offline' (不分大小寫)，
    則將該列第1到6欄的儲存格底色設為黃色 (色碼 F9F900)
    """
    yellow_fill = PatternFill(start_color="F9F900", end_color="F9F900", fill_type="solid")
    # 從第二列開始（假設第一列為標題）
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        cell_val = ws.cell(row=row[0].row, column=3).value
        if cell_val and isinstance(cell_val, str) and 'offline' in cell_val.lower():
            for col in range(1, 7):
                ws.cell(row=row[0].row, column=col).fill = yellow_fill

def beautify_excel(file_path):
    """
    開啟 _final.xlsx 檔案後，
    依序對每個工作表執行：
      1. 檢查每一列，若該列第3欄含 "offline"，則將該列第1到6欄的底色變為黃色
      2. 自動調整欄寬及列高 (模擬 Excel 全選後之格式調整)
    最後另存為 _beautified.xlsx
    """
    wb = load_workbook(file_path)
    for ws in wb.worksheets:
        highlight_offline_rows(ws)
        auto_adjust(ws)
    new_file_path = file_path.replace('_final.xlsx', '_beautified.xlsx')
    wb.save(new_file_path)
    print(f"Beautified file saved as {new_file_path}")

# ===== 主程序 =====

# 第一階段：處理所有原始 .xlsx 檔案 (排除已處理的 _final 與 _beautified 檔)
for file_name in os.listdir(folder_path):
    if file_name.endswith('.xlsx') and not file_name.endswith('_final.xlsx') and not file_name.endswith('_beautified.xlsx'):
        file_path = os.path.join(folder_path, file_name)
        process_excel(file_path)

# 第二階段：針對每個 _final.xlsx 檔案進行版面美化 (依據需求模擬 CTL+A→格式→自動調整)
for file_name in os.listdir(folder_path):
    if file_name.endswith('_final.xlsx'):
        file_path = os.path.join(folder_path, file_name)
        beautify_excel(file_path)





# ------------------------------------------------------------------------------------------
# 下面逐步說明此程式碼的各個部分與功能：

# ---

# ### 1. 匯入模組與設定共用樣式

# - **匯入模組：**  
#   程式碼先匯入 `os`、`datetime`、`timedelta`、`load_workbook` 以及 `openpyxl` 相關的樣式物件。這些模組用來操作檔案、處理日期時間以及修改 Excel 儲存格的格式。

# - **設定資料夾路徑：**  
#   變數 `folder_path` 指定了 Excel 檔案所在的資料夾路徑，請確保該路徑存在。

# - **定義共用字型、對齊方式與邊框樣式：**  
#   使用 `Font`、`Alignment` 與 `Border`（搭配 `Side`）設定統一的字型（新細明體，12 號）、置中對齊以及細線邊框，這些樣式稍後會套用到所有工作表與儲存格上。

# ---

# ### 2. 工具函式

# #### a. `utc_to_taiwan(utc_str)`

# - **功能：**  
#   將 UTC 格式的時間字串轉換為台灣時間（UTC+8）。
  
# - **實作細節：**  
#   先嘗試以包含毫秒的格式解析字串，若失敗則用不包含毫秒的格式。最後將 UTC 時間加上 8 小時後再轉換成字串格式回傳。

# #### b. `auto_adjust(ws)`

# - **功能：**  
#   自動調整工作表中每一欄的欄寬與每一列的列高，模擬 Excel 中「全選」後使用自動調整功能。

# - **實作細節：**  
#   - **欄寬調整：** 對每一欄遍歷所有儲存格，找出字串長度最大的值，並設定該欄寬為該長度加 2。  
#   - **列高調整：** 對每一列計算各儲存格換行後的行數（或文字行數），並根據最大行數來設定該列高度。

# #### c. `apply_formatting(ws)`

# - **功能：**  
#   對整個工作表的所有儲存格套用統一的字型、置中、邊框，然後調用 `auto_adjust` 自動調整欄寬與列高。

# ---

# ### 3. 主資料處理流程：`process_excel(file_path)`

# 這個函式負責從原始 Excel 檔案建立處理後的新檔案，主要步驟如下：

# 1. **讀取原始工作表：**  
#    從檔案中尋找名稱為 "Sheet1" 或 "sheet1" 的工作表作為原始資料來源。

# 2. **建立新工作表 sheet2：**  
#    複製原始資料到新工作表 `sheet2`，並將第4欄（UTC 時間）轉換為台灣時間（透過 `utc_to_taiwan` 函式）。

# 3. **根據 MAC address 分工作表：**  
#    以 `sheet2` 的第一欄（MAC address）作為依據，將資料分類後建立新的工作表（工作表名稱即為 MAC address），每個新工作表先複製標題列，再寫入對應的資料列。

# 4. **計算時間差與統計資料：**  
#    對除原始工作表（sheet1、sheet2）及 summary 外的其他工作表進行以下處理：
#    - 在第5欄計算每筆資料與上一筆資料的時間差（以秒為單位）。
#    - 統計第3欄中包含 "offline" 與 "connected" 的次數。
#    - 在第6欄（標題為 "斷線秒數"）中，若前一列的第3欄含 "offline"，則複製當列第5欄的數值。

# 5. **建立 summary 工作表：**  
#    將各 MAC 工作表的統計資料（例如最高斷線秒數、斷線次數、連線次數）彙總至一個名為 "summary" 的工作表中。

# 6. **重新命名原始工作表：**  
#    將原本的 sheet1 改名為 "Log(UTC)"，sheet2 改名為 "Log(TW)"。

# 7. **格式化所有工作表：**  
#    對所有工作表套用 `apply_formatting` 函式，統一字型、對齊、邊框與自動調整欄寬與列高。

# 8. **儲存處理結果：**  
#    最後將檔案儲存為檔名加上 `_final.xlsx` 的檔案。

# ---

# ### 4. 版面美化流程

# #### a. `highlight_offline_rows(ws)`

# - **功能：**  
#   遍歷單一工作表的每一列，檢查第3欄儲存格是否包含 "offline"（不區分大小寫）；若符合則將該列第1到6欄的儲存格背景設為黃色（色碼 F9F900）。

# #### b. `beautify_excel(file_path)`

# - **功能：**  
#   針對先前產生的 `_final.xlsx` 檔案，對每個工作表執行下列步驟：
#   1. 呼叫 `highlight_offline_rows`，檢查並標記 "offline" 的列。
#   2. 呼叫 `apply_formatting` 或 `auto_adjust` 來重新調整所有儲存格的格式，模擬 Excel 中全選後的格式自動調整（自動調整欄寬與列高）。
# - **儲存：**  
#   美化完成後，將檔案另存為 `_beautified.xlsx`。

# ---

# ### 5. 主程序流程

# 程式碼最後分為兩個階段進行檔案處理：

# 1. **第一階段：處理原始檔案**  
#    - 遍歷 `folder_path` 中所有 `.xlsx` 檔案（排除檔名已包含 `_final.xlsx` 與 `_beautified.xlsx` 的檔案）。
#    - 對每個原始檔案執行 `process_excel`，產生處理後的 `_final.xlsx` 檔案。

# 2. **第二階段：版面美化**  
#    - 遍歷資料夾中所有以 `_final.xlsx` 結尾的檔案。
#    - 對每個檔案執行 `beautify_excel`，將最終結果另存為 `_beautified.xlsx`。

# ---

# ### 總結

# - **程式功能：**  
#   此程式先將原始 Excel 資料依照特定邏輯處理、分工作表、計算統計數據並建立 summary，然後再針對結果進行版面美化（包含自動調整欄寬、列高，及針對 "offline" 狀態標記背景色）。

# - **實作方式：**  
#   利用 openpyxl 讀取與修改 Excel 檔案，分別呼叫各個函式來達成資料處理與版面美化的效果。整個流程分成兩個階段：先處理產生 _final.xlsx，再以 _final.xlsx 為基礎進行版面美化產生 _beautified.xlsx。

# 以上就是對此程式碼各部分功能與流程的詳細說明！
