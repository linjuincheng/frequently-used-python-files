# 數個excel檔, 每個excel裡面有都有工作表名稱為缺漏筆數, 這工作表的第1欄為Start Time, 第2欄為End Time, 如果要在這些excel檔裡的缺漏筆數工作表產出第3欄(名為:time difference), 第3欄為第2欄與第1欄的時間秒數落差, 此外, 輸出的excel檔名要和原始檔名一樣(把原始檔名覆蓋過去),Python怎麼寫?
# 你可以使用 pandas 和 openpyxl 來處理這個任務。以下是處理數個 Excel 檔案的 Python 程式碼，它會在每個 Excel 檔案中名為 缺漏筆數 的工作表中添加一個新的欄位 time difference，計算 Start Time 和 End Time 之間的秒數差異，並將結果保存回原始檔案。
# 程式碼說明
# 導入必要的庫：

# pandas 用於數據處理和 DataFrame 操作。
# openpyxl 用於讀取和寫入 Excel 文件。
# 設定目錄：

# input_directory 指定了需要處理的目錄。
# 定義 process_excel 函數：

# 使用 load_workbook 讀取 Excel 文件。
# 檢查是否存在名為 缺漏筆數 的工作表。
# 將工作表內容讀取到 DataFrame 中，並設置列名。
# 將 Start Time 和 End Time 轉換為 datetime 格式。
# 計算 Start Time 和 End Time 之間的秒數差異，並將結果寫入新的 time difference 欄位。
# 使用 pd.ExcelWriter 將修改後的 DataFrame 保存回原始檔案，覆蓋原來的工作表。
# 遍歷目錄中的 Excel 文件：

# 使用 os.listdir 列出目錄中的所有文件，對每個 .xlsx 文件調用 process_excel 函數進行處理。
# 注意事項
# 備份原始檔案：在覆蓋原始檔案之前，建議先備份原始檔案，以防出現意外情況。
# 安裝庫：確保已安裝 pandas 和 openpyxl 庫。如果尚未安裝，可以使用 pip install pandas openpyxl 安裝。
# 這樣，你就能夠處理所有 Excel 檔案，並在每個檔案的 缺漏筆數 工作表中添加 time difference 欄位。



import os
import pandas as pd
from openpyxl import load_workbook

# 設定需要處理的目錄
input_directory = 'C:\\_python 2024\\2024final'

# 定義處理 Excel 文件的函數
def process_excel(file_path):
    # 讀取工作簿
    wb = load_workbook(file_path)
    
    # 檢查是否存在名為 '缺漏筆數' 的工作表
    if '缺漏筆數' in wb.sheetnames:
        ws = wb['缺漏筆數']

        # 將工作表內容讀取到 DataFrame
        df = pd.DataFrame(ws.values)
        
        # 設定列名
        df.columns = df.iloc[0]
        df = df[1:]

        # 確保時間欄位是日期時間格式
        df['Start Time'] = pd.to_datetime(df['Start Time'])
        df['End Time'] = pd.to_datetime(df['End Time'])

        # 計算時間差（秒數）
        df['time difference'] = (df['End Time'] - df['Start Time']).dt.total_seconds()

        # 將結果寫回到原 Excel 檔案
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name='缺漏筆數', index=False)

# 遍歷目錄中的所有 Excel 文件並處理
for filename in os.listdir(input_directory):
    if filename.endswith('.xlsx'):
        file_path = os.path.join(input_directory, filename)
        process_excel(file_path)

